import pandas as pd, re, unicodedata
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys, os

# ── Parámetros ─────────────────────────────────────────────
csv_path   = sys.argv[1]
n_sub      = int(sys.argv[2])
output_dir = sys.argv[3]

# ── Naming Cabida_N ────────────────────────────────────────
counter = 1
while True:
    output_name = f"Cabida_{counter}.xlsx"
    output_path = os.path.join(output_dir, output_name)
    if not os.path.exists(output_path): break
    counter += 1

# ── Colores canónicos por función ─────────────────────────
COLOR_CANONICO = {
    'Residencial Util':    '2E74B5',
    'Residencial Comun':   '1F3864',
    'Residencial Terraza': '9DC3E6',
    'Residencial Loggia':  'C5D9F1',
    'Comercial Util':      'F4802A',
    'Comercial Comun':     '843C0C',
    'Oficinas Util':       '70AD47',
    'Oficinas Comun':      '375623',
    'Estacionamientos':    'E2EFDA',
    'Ascensores':          'BFBFBF',
    'Otro':                'F2F2F2',
}
COLOR_TINT = {
    'Residencial Util':    'D6E4F4',
    'Residencial Comun':   'D0D9E8',
    'Residencial Terraza': 'EBF4FB',
    'Residencial Loggia':  'F4F8FD',
    'Comercial Util':      'FDE8D8',
    'Comercial Comun':     'E8D5CD',
    'Oficinas Util':       'E2F0DA',
    'Oficinas Comun':      'D5E2CF',
    'Estacionamientos':    'F5FAF5',
    'Ascensores':          'F5F5F5',
    'Otro':                'FAFAFA',
}
FT_BLANCO  = 'FFFFFF'
FT_OSCURO  = '1A1A1A'
HDR_OSCUROS = {'Residencial Util','Residencial Comun','Comercial Comun','Oficinas Comun'}

C_HDR_BG   = '1F3864'
C_HDR_FT   = 'FFFFFF'
C_TOTAL_BG = '2E74B5'
C_TOTAL_FT = 'FFFFFF'
C_SUBTOT   = 'D6E4F0'
C_ROW_ODD  = 'F7FAFD'
C_ROW_EVEN = 'FFFFFF'
C_BORDER   = 'BDD7EE'
C_BORDER_H = '4472C4'

CANONICO = {
    'ResidencialÚtil':'Residencial Util','ResidencialUtil':'Residencial Util',
    'ResidencialComún':'Residencial Comun','ResidencialComun':'Residencial Comun',
    'ResidencialTerraza':'Residencial Terraza','ResidencialLoggia':'Residencial Loggia',
    'ComercialÚtil':'Comercial Util','ComercialUtil':'Comercial Util',
    'ComercialComún':'Comercial Comun','ComercialComun':'Comercial Comun',
    'OficinasÚtil':'Oficinas Util','OficinasUtil':'Oficinas Util',
    'OficinasComún':'Oficinas Comun','OficinasComun':'Oficinas Comun',
    'Estacionamientos':'Estacionamientos','Parking':'Estacionamientos',
    'CirculaciónVertical':'Ascensores','CirculacionVertical':'Ascensores',
    'Circulación Vertical':'Ascensores','Ascensores':'Ascensores','Core':'Ascensores',
}
FACTOR_VENTA = {
    'Residencial Util':1.0,'Residencial Comun':0.0,'Residencial Terraza':0.5,
    'Residencial Loggia':1.0,'Comercial Util':1.0,'Comercial Comun':0.0,
    'Oficinas Util':1.0,'Oficinas Comun':0.0,'Estacionamientos':0.0,
    'Ascensores':0.0,'Otro':0.0,
}

# Mapeo GFA -> Tipologia (Residencial Util)
# Editable segun proyecto. Tolerancia +/- 1 m2 por redondeo en Forma.
TIPOLOGIA_RANGOS = [
    # (gfa_min, gfa_max, nombre_tipologia, dorm, banos, programa)
    (40, 50, '1D1B', 1, 1, '1 dorm + 1 bano'),
    (60, 80, '2D2B', 2, 2, '2 dorm + 2 banos'),
    (81, 95, '3D2B', 3, 2, '3 dorm + 2 banos'),
    (96, 120, '3D2B-XL', 3, 2, '3 dorm + 2 banos (XL)'),
]
PROGRAMA_TIPOLOGIA = {t: prog for _,_,t,_,_,prog in TIPOLOGIA_RANGOS}
COLOR_TIPOLOGIA = {
    '1D1B':    'B4C7E7',
    '2D2B':    '8FAADC',
    '3D2B':    '2E74B5',
    '3D2B-XL': '1F3864',
    'Otro':    'BFBFBF',
}

def tipologia(gfa, canonico):
    '''Devuelve la tipologia segun GFA. Solo aplica a Residencial Util.'''
    if canonico != 'Residencial Util': return ''
    g = float(gfa) if gfa else 0
    # Excluir unidades sin volumen (areas remanentes en Forma)
    if g <= 0: return ''
    for gmin, gmax, nombre, _, _, _ in TIPOLOGIA_RANGOS:
        if gmin <= g <= gmax: return nombre
    return f'Otro ({g:.0f}m2)'
CONSTRUIDO = {k:True for k in FACTOR_VENTA}
CONSTRUIDO['Ascensores'] = False
CONSTRUIDO['Otro']       = False

ORDEN_FUNC = [
    'Residencial Util','Residencial Comun','Residencial Terraza','Residencial Loggia',
    'Comercial Util','Comercial Comun','Oficinas Util','Oficinas Comun',
    'Estacionamientos','Ascensores','Otro'
]
GRUPOS = {
    'Residencial': ['Residencial Util','Residencial Comun','Residencial Terraza','Residencial Loggia'],
    'Comercial':   ['Comercial Util','Comercial Comun'],
    'Oficinas':    ['Oficinas Util','Oficinas Comun'],
    'Otros usos':  ['Estacionamientos','Ascensores','Otro'],
}

def canon(f):
    v = unicodedata.normalize('NFC', str(f).strip())
    if v in CANONICO: return CANONICO[v]
    # Intentar tambien sin espacios internos (Forma a veces exporta "Residencial Util" con espacio)
    v_compact = v.replace(' ', '')
    if v_compact in CANONICO: return CANONICO[v_compact]
    vs = ''.join(c for c in unicodedata.normalize('NFD',v) if unicodedata.category(c)!='Mn')
    vs_compact = vs.replace(' ', '')
    for k,c2 in CANONICO.items():
        ks = ''.join(ch for ch in unicodedata.normalize('NFD',k) if unicodedata.category(ch)!='Mn')
        if vs.lower()==ks.lower(): return c2
        if vs_compact.lower()==ks.lower(): return c2
    return 'Otro'

def extraer_nivel(s):
    m = re.search(r'/(\d+)_', str(s))
    if m: return int(m.group(1))
    return None

def etiqueta_fn(x):
    if x is None: return 'S/N'
    if x < n_sub: return f'ST-{n_sub - x}'
    return f'N{x - n_sub + 1}'

def limpiar_gfa(v):
    s = str(v).replace('\xa0','').replace('\u202f','').replace(' ','')
    return pd.to_numeric(s, errors='coerce')

def fill(c):  return PatternFill('solid', fgColor=c)
def fnt(bold=False, sz=10, color='1A1A1A', name='Arial Narrow'):
    return Font(name=name, bold=bold, size=sz, color=color)
def brd(color=C_BORDER):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def aln(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def hdr(ws, r, c, val, bg=C_HDR_BG, ft=C_HDR_FT, sz=10, bold=True, wrap=False, ha='center'):
    cell = ws.cell(r, c, val)
    cell.fill=fill(bg); cell.font=fnt(bold,sz,ft)
    cell.alignment=aln(ha,wrap=wrap); cell.border=brd(C_BORDER_H)
    return cell

def tot(ws, r, c, val, fmt=None, ha='right'):
    cell = ws.cell(r, c, val)
    cell.fill=fill(C_TOTAL_BG); cell.font=fnt(True,10,C_TOTAL_FT)
    cell.alignment=aln(ha); cell.border=brd(C_HDR_BG)
    if fmt: cell.number_format=fmt
    return cell

# ── Leer CSV ──────────────────────────────────────────────
for enc in ['utf-8','utf-8-sig','latin-1','cp1252']:
    try:
        with open(csv_path, encoding=enc) as f: raw_text = f.read()
        # Normalizar terminadores de linea (\r solo -> \n)
        raw_text = raw_text.replace('\r\n', '\n').replace('\r', '\n')
        import io
        line = raw_text.split('\n')[0]
        sep = ';' if line.count(';') > line.count(',') else ','
        df  = pd.read_csv(io.StringIO(raw_text), sep=sep)
        # Normalizar headers ES->EN (Forma puede exportar en espanol)
        header_map = {
            'ID': 'Id', 'id': 'Id',
            'Tipo': 'Type', 'tipo': 'Type',
            'Funci\u00f3n': 'Function', 'Funcion': 'Function',
            'funci\u00f3n': 'Function', 'funcion': 'Function',
        }
        df.rename(columns=header_map, inplace=True)
        break
    except: continue

df['GFA']       = df['GFA'].apply(limpiar_gfa).fillna(0)
df['nivel_raw'] = df['Id'].apply(extraer_nivel)
df['Etiqueta']  = df['nivel_raw'].apply(etiqueta_fn)
df['Canonico']  = df['Function'].apply(canon)
df['FV']        = df['Canonico'].map(FACTOR_VENTA).fillna(0)
df['SV']        = df['GFA'] * df['FV']
df['Integra']   = df['Canonico'].map(CONSTRUIDO).fillna(False)

# === Tipologias (Residencial Util) y reparto de terraza por unidad por nivel ===
df['building'] = df['Id'].astype(str).str.extract(r'root/([^/]+)/')
df['Tipologia'] = df.apply(lambda r: tipologia(r['GFA'], r['Canonico']), axis=1)

# Calcular terraza prorrateada: total terraza por (building,nivel) / num unidades RU en (building,nivel)
ru_mask = df['Canonico']=='Residencial Util'
ter_mask = df['Canonico']=='Residencial Terraza'

# Reparto PROPORCIONAL: terraza total del proyecto se distribuye entre unidades RU
# segun el GFA util de cada una. Asi una unidad mas grande recibe mas terraza.
terraza_total_proyecto = df[ter_mask]['GFA'].sum()
gfa_util_total_proyecto = df[ru_mask]['GFA'].sum()
factor_terraza = terraza_total_proyecto / gfa_util_total_proyecto if gfa_util_total_proyecto > 0 else 0

def terraza_unidad(row):
    if row['Canonico']!='Residencial Util': return 0.0
    return round(row['GFA'] * factor_terraza, 2)

df['Terraza_Unidad'] = df.apply(terraza_unidad, axis=1)
df['GFA_Total_Unidad'] = df['GFA'] + df['Terraza_Unidad']  # util + terraza prorrateada
df['SV_Unidad'] = df['GFA'] + 0.5 * df['Terraza_Unidad']   # venta = util*1.0 + terraza*0.5

funciones_presentes = [f for f in ORDEN_FUNC if f in df['Canonico'].unique()]

def sort_etq(e):
    if e.startswith('ST-'):
        try: return -1000 + int(e[3:])
        except: return -999
    elif e.startswith('N'):
        try: return int(e[1:])
        except: return 9999
    return 9998

etiquetas = sorted(df['Etiqueta'].unique().tolist(), key=sort_etq)
df_sorted = df.sort_values(['nivel_raw','Canonico']).reset_index(drop=True)
n_data    = len(df_sorted)

wb = Workbook()

# ═══════════════════════════════════════════════════════════
# HOJA 1 — DATOS
# A=Id  B=Type  C=Function(raw)  D=GFA  E=Nivel  F=Etiqueta
# G=Función Canónica  H=Factor Venta  I=Sup.Venta  J=Integra
# Datos desde fila 3  (fila 1=título, fila 2=headers)
# ═══════════════════════════════════════════════════════════
ws_d = wb.active
ws_d.title = 'Datos'

ws_d.merge_cells('A1:J1')
c=ws_d.cell(1,1,'DATOS — DETALLE POR ELEMENTO  |  Fuente de Cabida y Tipología')
c.fill=fill(C_HDR_BG); c.font=fnt(True,13,C_HDR_FT); c.alignment=aln('center')
ws_d.row_dimensions[1].height=28

DATOS_HDRS  = ['Id','Type','Function (raw)','GFA m²','Nivel','Etiqueta',
               'Función Canónica','Factor Venta','Sup. Venta m²','Integra Construido']
DATOS_WIDTHS= [42,14,24,12,8,10,22,14,16,18]
for i,(h,w) in enumerate(zip(DATOS_HDRS,DATOS_WIDTHS),1):
    hdr(ws_d,2,i,h); ws_d.column_dimensions[get_column_letter(i)].width=w
ws_d.row_dimensions[2].height=20

for r_i, row in enumerate(df_sorted.itertuples(index=False), 3):
    is_sub = str(row.Etiqueta).startswith('ST-')
    bg = 'D6E4F0' if is_sub else (C_ROW_ODD if r_i%2==0 else C_ROW_EVEN)
    bld = is_sub
    vals=[
        (row.Id,        '@',          'left'),
        (getattr(row,'Type',''), '@', 'left'),
        (row.Function,  '@',          'left'),
        (row.GFA,       '#,##0.00',   'right'),
        (row.nivel_raw, '0',          'center'),
        (row.Etiqueta,  '@',          'center'),
        (row.Canonico,  '@',          'left'),
        (row.FV,        '0%',         'center'),
        (row.SV,        '#,##0.00',   'right'),
        ('✓' if row.Integra else '—','@','center'),
    ]
    for c_i,(val,fmt,ha) in enumerate(vals,1):
        cell=ws_d.cell(r_i,c_i,val)
        cell.fill=fill(bg); cell.font=fnt(bld,9)
        cell.alignment=aln(ha); cell.border=brd(); cell.number_format=fmt
    ws_d.row_dimensions[r_i].height=16

TR = 3 + n_data   # fila total Datos
tot(ws_d,TR,1,'TOTAL',ha='center')
for ci in range(2,11): tot(ws_d,TR,ci,None)
tot(ws_d,TR,4, df_sorted['GFA'].sum(),'#,##0.00')
tot(ws_d,TR,9, df_sorted['SV'].sum(), '#,##0.00')
ws_d.row_dimensions[TR].height=20
ws_d.freeze_panes='A3'

# Rangos de referencia para SUMIFS (sin fila de total)
D_GFA = f"Datos!$D$3:$D${TR-1}"
D_ETQ = f"Datos!$F$3:$F${TR-1}"
D_CAN = f"Datos!$G$3:$G${TR-1}"
D_SV  = f"Datos!$I$3:$I${TR-1}"

# ═══════════════════════════════════════════════════════════
# HOJA 2 — CABIDA  (todas las celdas de datos = SUMIFS → Datos)
# Fila 1=título  Fila 2=grupos  Fila 3=funciones (coloreadas)
# Filas 4..N=etiquetas  Fila N+1=TOTAL  Filas N+3..=KPIs
# ═══════════════════════════════════════════════════════════
ws_c = wb.create_sheet('Cabida')

fn_cols = {}
col_ptr = 2
for fn in funciones_presentes:
    fn_cols[fn] = col_ptr; col_ptr += 1
COL_TOTAL = col_ptr; col_ptr += 1
COL_VENTA = col_ptr; col_ptr += 1
LAST_COL  = col_ptr - 1

ws_c.merge_cells(f'A1:{get_column_letter(LAST_COL)}1')
c=ws_c.cell(1,1,'CABIDA NUMÉRICA — SUPERFICIES POR NIVEL')
c.fill=fill(C_HDR_BG); c.font=fnt(True,13,C_HDR_FT)
c.alignment=aln('center'); ws_c.row_dimensions[1].height=28

# Fila 2 — grupos
col_g = 2
for grp, fns in GRUPOS.items():
    fns_p = [f for f in fns if f in funciones_presentes]
    if not fns_p: continue
    c1=col_g; c2=col_g+len(fns_p)-1
    if c2>c1: ws_c.merge_cells(f'{get_column_letter(c1)}2:{get_column_letter(c2)}2')
    cell=ws_c.cell(2,c1,grp)
    cell.fill=fill(C_TOTAL_BG); cell.font=fnt(True,9,C_HDR_FT)
    cell.alignment=aln('center'); cell.border=brd(C_BORDER_H)
    col_g += len(fns_p)
for lbl,col in [('TOTAL m²',COL_TOTAL),('VENTA m²',COL_VENTA)]:
    cell=ws_c.cell(2,col,lbl)
    cell.fill=fill(C_TOTAL_BG); cell.font=fnt(True,9,C_HDR_FT)
    cell.alignment=aln('center'); cell.border=brd(C_BORDER_H)

# Fila 3 — "Nivel" + nombres función coloreados
ws_c.merge_cells('A2:A3')
c=ws_c.cell(2,1,'Nivel')
c.fill=fill(C_HDR_BG); c.font=fnt(True,10,C_HDR_FT)
c.alignment=aln('center','center'); c.border=brd(C_BORDER_H)
ws_c.column_dimensions['A'].width=10

for fn, col in fn_cols.items():
    bg  = COLOR_CANONICO.get(fn,'BFBFBF')
    ft  = FT_BLANCO if fn in HDR_OSCUROS else FT_OSCURO
    cell=ws_c.cell(3,col,fn)
    cell.fill=fill(bg); cell.font=fnt(True,8,ft)
    cell.alignment=aln('center',wrap=True); cell.border=brd(C_BORDER_H)
    ws_c.column_dimensions[get_column_letter(col)].width=15

for col,lbl in [(COL_TOTAL,'Total m²'),(COL_VENTA,'Venta m²')]:
    cell=ws_c.cell(3,col,lbl)
    cell.fill=fill(C_TOTAL_BG); cell.font=fnt(True,9,C_HDR_FT)
    cell.alignment=aln('center'); cell.border=brd(C_BORDER_H)
    ws_c.column_dimensions[get_column_letter(col)].width=13
ws_c.row_dimensions[2].height=22; ws_c.row_dimensions[3].height=42

# Filas de datos — SUMIFS vinculados a Datos
DS = 4   # DATA_START en Cabida
for r_i, etq in enumerate(etiquetas):
    row = DS + r_i
    is_sub = etq.startswith('ST-')
    bg_lbl = 'D6E4F0' if is_sub else (C_ROW_ODD if r_i%2==0 else C_ROW_EVEN)
    cell=ws_c.cell(row,1,etq)
    cell.fill=fill(bg_lbl); cell.font=fnt(is_sub,10)
    cell.alignment=aln('center'); cell.border=brd()

    etq_ref = f'$A{row}'
    construido_col_ltrs = []
    for fn, col in fn_cols.items():
        fn_ref = f'{get_column_letter(col)}$3'
        formula = f'=IFERROR(SUMIFS({D_GFA},{D_ETQ},{etq_ref},{D_CAN},{fn_ref}),0)'
        tint = COLOR_TINT.get(fn, C_ROW_ODD)
        cell2=ws_c.cell(row,col,formula)
        cell2.fill=fill(tint); cell2.font=fnt(False,10)
        cell2.alignment=aln('right'); cell2.border=brd(); cell2.number_format='#,##0'
        if CONSTRUIDO.get(fn,False): construido_col_ltrs.append(get_column_letter(col))

    # Total m² = suma de columnas construido en esa fila (referenciadas)
    f_tot = '='+'+'.join(f'{c}{row}' for c in construido_col_ltrs) if construido_col_ltrs else '=0'
    c_tot=ws_c.cell(row,COL_TOTAL,f_tot)
    c_tot.fill=fill(C_SUBTOT); c_tot.font=fnt(True,10)
    c_tot.alignment=aln('right'); c_tot.border=brd(C_HDR_BG); c_tot.number_format='#,##0'

    # Venta m² = SUMIFS Sup.Venta en Datos por etiqueta
    c_sv=ws_c.cell(row,COL_VENTA,f'=IFERROR(SUMIFS({D_SV},{D_ETQ},{etq_ref}),0)')
    c_sv.fill=fill(C_SUBTOT); c_sv.font=fnt(True,10)
    c_sv.alignment=aln('right'); c_sv.border=brd(C_HDR_BG); c_sv.number_format='#,##0'
    ws_c.row_dimensions[row].height=18

# Fila TOTAL — suma de cada columna de datos
TOT_C = DS + len(etiquetas)
tot(ws_c,TOT_C,1,'TOTAL',ha='center')
for fn, col in fn_cols.items():
    ltr=get_column_letter(col)
    c=ws_c.cell(TOT_C,col,f'=SUM({ltr}{DS}:{ltr}{TOT_C-1})')
    c.fill=fill(C_TOTAL_BG); c.font=fnt(True,10,C_TOTAL_FT)
    c.alignment=aln('right'); c.border=brd(C_HDR_BG); c.number_format='#,##0'
for col in [COL_TOTAL,COL_VENTA]:
    ltr=get_column_letter(col)
    c=ws_c.cell(TOT_C,col,f'=SUM({ltr}{DS}:{ltr}{TOT_C-1})')
    c.fill=fill(C_TOTAL_BG); c.font=fnt(True,10,C_TOTAL_FT)
    c.alignment=aln('right'); c.border=brd(C_HDR_BG); c.number_format='#,##0'
ws_c.row_dimensions[TOT_C].height=20

# KPIs referenciados a celdas de Cabida
KR = TOT_C + 2
venta_ref = f'{get_column_letter(COL_VENTA)}{TOT_C}'
tot_ref   = f'{get_column_letter(COL_TOTAL)}{TOT_C}'
efic_cell = f'B{KR+2}'
for r,(lbl,formula,fmt) in enumerate([
    ('Sup. Venta',      f'={venta_ref}',                     '#,##0" m²"'),
    ('Sup. Construida', f'={tot_ref}',                       '#,##0" m²"'),
    ('Eficiencia',      f'=IFERROR({venta_ref}/{tot_ref},0)', '0.0%'),
    ('Calificación',    f'=IF({efic_cell}>0.75,"🟢 Alta",IF({efic_cell}>0.5,"🟡 Media","🔴 Baja"))','@'),
], KR):
    c1=ws_c.cell(r,1,lbl)
    c1.fill=fill('1F4E79'); c1.font=fnt(True,10,C_HDR_FT)
    c1.alignment=aln('left'); c1.border=brd(C_BORDER_H)
    c2=ws_c.cell(r,2,formula)
    c2.fill=fill(C_SUBTOT); c2.font=fnt(True,10)
    c2.alignment=aln('right'); c2.border=brd(); c2.number_format=fmt

ws_c.freeze_panes='B4'

# ═══════════════════════════════════════════════════════════
# HOJA 3 — PROGRAMA  (SUMIFS → Datos)
# A=Función  B=GFA  C=Venta  D=FV  E=%Construido  F=Integra  G=Color
# Datos desde fila 3
# ═══════════════════════════════════════════════════════════
ws_t = wb.create_sheet('Programa')

ws_t.merge_cells('A1:G1')
c=ws_t.cell(1,1,'PROGRAMA — RESUMEN POR FUNCIÓN CANÓNICA')
c.fill=fill(C_HDR_BG); c.font=fnt(True,13,C_HDR_FT); c.alignment=aln('center')
ws_t.row_dimensions[1].height=28

T_HDRS  = ['Función Canónica','GFA Total m²','Sup. Venta m²','Factor Venta','% Construido','Integra Construido','Color']
T_WIDTHS= [24,16,16,14,14,18,12]
for i,(h,w) in enumerate(zip(T_HDRS,T_WIDTHS),1):
    hdr(ws_t,2,i,h); ws_t.column_dimensions[get_column_letter(i)].width=w
ws_t.row_dimensions[2].height=20

fn_row_map = {fn: 3+i for i,fn in enumerate(funciones_presentes)}
construido_rows = [fn_row_map[fn] for fn in funciones_presentes if CONSTRUIDO.get(fn,False)]
tot_construido_f = '+'.join(f'B{r}' for r in construido_rows) if construido_rows else '0'

for fn in funciones_presentes:
    r   = fn_row_map[fn]
    odd = r % 2 == 0
    bg_fn = COLOR_CANONICO.get(fn,'BFBFBF')
    ft_fn = FT_BLANCO if fn in HDR_OSCUROS else FT_OSCURO
    integra = CONSTRUIDO.get(fn,False)
    fv      = FACTOR_VENTA.get(fn,0)
    bg_d    = C_ROW_ODD if odd else C_ROW_EVEN
    fn_ref  = f'$A{r}'

    # A — nombre con color canónico
    c=ws_t.cell(r,1,fn)
    c.fill=fill(bg_fn); c.font=fnt(True,10,ft_fn)
    c.alignment=aln('left'); c.border=brd(C_BORDER_H)

    # B — GFA SUMIFS → Datos!G (Función Canónica)
    c=ws_t.cell(r,2,f'=IFERROR(SUMIFS({D_GFA},{D_CAN},{fn_ref}),0)')
    c.fill=fill(bg_d); c.font=fnt(False,10)
    c.alignment=aln('right'); c.border=brd(); c.number_format='#,##0'

    # C — Sup.Venta SUMIFS → Datos!I
    c=ws_t.cell(r,3,f'=IFERROR(SUMIFS({D_SV},{D_CAN},{fn_ref}),0)')
    c.fill=fill(bg_d); c.font=fnt(False,10)
    c.alignment=aln('right'); c.border=brd(); c.number_format='#,##0'

    # D — Factor Venta (estático del mapeo)
    c=ws_t.cell(r,4,fv)
    c.fill=fill(bg_d); c.font=fnt(False,10)
    c.alignment=aln('center'); c.border=brd(); c.number_format='0%'

    # E — % del total construido (referenciado a celdas B de esta hoja)
    if integra:
        c=ws_t.cell(r,5,f'=IFERROR(B{r}/({tot_construido_f}),0)')
        c.number_format='0.0%'
    else:
        c=ws_t.cell(r,5,'—')
        c.number_format='@'
    c.fill=fill(bg_d); c.font=fnt(False,10)
    c.alignment=aln('center'); c.border=brd()

    # F — Integra Construido
    c=ws_t.cell(r,6,'✓ Sí' if integra else '✗ No')
    c.fill=fill(bg_d); c.font=fnt(False,10)
    c.alignment=aln('center'); c.border=brd()

    # G — Swatch color canónico
    c=ws_t.cell(r,7,f'#{bg_fn}')
    c.fill=fill(bg_fn); c.font=fnt(False,9,ft_fn)
    c.alignment=aln('center'); c.border=brd()
    ws_t.row_dimensions[r].height=18

# Fila TOTAL — referencias a celdas B y C de esta hoja
TOT_T = 3 + len(funciones_presentes)
tot(ws_t,TOT_T,1,'TOTAL',ha='center')
gfa_f = '='+'+'.join(f'B{fn_row_map[fn]}' for fn in funciones_presentes)
sv_f  = '='+'+'.join(f'C{fn_row_map[fn]}' for fn in funciones_presentes)
c=ws_t.cell(TOT_T,2,gfa_f)
c.fill=fill(C_TOTAL_BG); c.font=fnt(True,10,C_TOTAL_FT)
c.alignment=aln('right'); c.border=brd(C_HDR_BG); c.number_format='#,##0'
c=ws_t.cell(TOT_T,3,sv_f)
c.fill=fill(C_TOTAL_BG); c.font=fnt(True,10,C_TOTAL_FT)
c.alignment=aln('right'); c.border=brd(C_HDR_BG); c.number_format='#,##0'
for col in [4,5,6,7]:
    c=ws_t.cell(TOT_T,col,None)
    c.fill=fill(C_TOTAL_BG); c.font=fnt(True,10,C_TOTAL_FT)
    c.alignment=aln('center'); c.border=brd(C_HDR_BG)
ws_t.row_dimensions[TOT_T].height=20
ws_t.freeze_panes='A3'


# ═════════════════════════════════════════════════════════════════════════════
# HOJA 4 — TIPOLOGIAS (resumen + pivot Nivel x Tipologia)
# A=Tipologia  B=N Unidades  C=GFA Util Prom  D=Terraza Prom
# E=Total Prom (Util+Terr)  F=Sup Venta Prom  G=GFA Util Total  H=SV Total
# ═════════════════════════════════════════════════════════════════════════════
ru_df = df[(df['Canonico']=='Residencial Util') & (df['GFA']>0)].copy()
tipologias_presentes = [t for _,_,t,_,_,_ in TIPOLOGIA_RANGOS if t in ru_df['Tipologia'].unique()]
otros_tipos = sorted([t for t in ru_df['Tipologia'].unique() if t not in [x[2] for x in TIPOLOGIA_RANGOS]])
tipologias_orden = tipologias_presentes + otros_tipos

ws_tip = wb.create_sheet('Tipologias')
ws_tip.merge_cells('A1:I1')
c=ws_tip.cell(1,1,'TIPOLOGIAS RESIDENCIALES — DETALLE POR UNIDAD Y NIVEL')
c.fill=fill(C_HDR_BG); c.font=fnt(True,13,C_HDR_FT); c.alignment=aln('center')
ws_tip.row_dimensions[1].height=28

# Tabla resumen por tipologia (filas 3-N)
T_HDRS = ['Tipologia','Programa','N Unidades','GFA Util Prom m2','Terraza Prom m2','Total Prom m2','Sup Venta Prom m2','GFA Util Total','Sup Venta Total']
T_W    = [12,22,12,16,16,16,16,16,16]
for i,(h,w) in enumerate(zip(T_HDRS,T_W),1):
    hdr(ws_tip,2,i,h); ws_tip.column_dimensions[get_column_letter(i)].width=w
ws_tip.row_dimensions[2].height=22

for r_i, tip in enumerate(tipologias_orden):
    r = 3 + r_i
    sub = ru_df[ru_df['Tipologia']==tip]
    n_un = len(sub)
    gfa_prom = sub['GFA'].mean() if n_un>0 else 0
    ter_prom = sub['Terraza_Unidad'].mean() if n_un>0 else 0
    tot_prom = sub['GFA_Total_Unidad'].mean() if n_un>0 else 0
    sv_prom  = sub['SV_Unidad'].mean() if n_un>0 else 0
    gfa_tot  = sub['GFA'].sum()
    sv_tot   = sub['SV_Unidad'].sum()
    bg_tip   = COLOR_TIPOLOGIA.get(tip,'BFBFBF')
    is_dark  = tip in ('3D2B','3D2B-XL')
    ft_tip   = FT_BLANCO if is_dark else FT_OSCURO

    c=ws_tip.cell(r,1,tip)
    c.fill=fill(bg_tip); c.font=fnt(True,10,ft_tip)
    c.alignment=aln('center'); c.border=brd(C_BORDER_H)

    odd = r % 2 == 0
    bg_d = C_ROW_ODD if odd else C_ROW_EVEN
    programa = PROGRAMA_TIPOLOGIA.get(tip, '-')

    # Col B - Programa (texto descriptivo)
    c=ws_tip.cell(r,2,programa)
    c.fill=fill(bg_d); c.font=fnt(False,10)
    c.alignment=aln('left'); c.border=brd()

    vals = [
        (n_un,    '0',         'center'),
        (gfa_prom,'#,##0.0',   'right'),
        (ter_prom,'#,##0.0',   'right'),
        (tot_prom,'#,##0.0',   'right'),
        (sv_prom, '#,##0.0',   'right'),
        (gfa_tot, '#,##0',     'right'),
        (sv_tot,  '#,##0',     'right'),
    ]
    for c_i,(v,fmt_,ha) in enumerate(vals,3):
        c=ws_tip.cell(r,c_i,v)
        c.fill=fill(bg_d); c.font=fnt(False,10)
        c.alignment=aln(ha); c.border=brd(); c.number_format=fmt_
    ws_tip.row_dimensions[r].height=18

TOT_R = 3 + len(tipologias_orden)
tot(ws_tip,TOT_R,1,'TOTAL',ha='center')
tot_n   = len(ru_df)
tot_gfa = ru_df['GFA'].sum()
tot_sv  = ru_df['SV_Unidad'].sum()
tot(ws_tip,TOT_R,2,None)              # col Programa vacia en TOTAL
tot(ws_tip,TOT_R,3,tot_n,'0','center')
for ci in [4,5,6,7]: tot(ws_tip,TOT_R,ci,None)
tot(ws_tip,TOT_R,8,tot_gfa,'#,##0','right')
tot(ws_tip,TOT_R,9,tot_sv,'#,##0','right')
ws_tip.row_dimensions[TOT_R].height=20

# === Pivot Nivel x Tipologia (conteo unidades) ===
PIV_R = TOT_R + 3
ws_tip.merge_cells(start_row=PIV_R, start_column=1, end_row=PIV_R, end_column=len(tipologias_orden)+2)
c=ws_tip.cell(PIV_R,1,'CONTEO DE UNIDADES POR TIPOLOGIA Y NIVEL')
c.fill=fill(C_TOTAL_BG); c.font=fnt(True,11,C_TOTAL_FT); c.alignment=aln('center')
ws_tip.row_dimensions[PIV_R].height=22

hdr(ws_tip, PIV_R+1, 1, 'Nivel')
for i,tip in enumerate(tipologias_orden):
    bg_tip = COLOR_TIPOLOGIA.get(tip,'BFBFBF')
    is_dark = tip in ('3D2B','3D2B-XL')
    ft_tip = FT_BLANCO if is_dark else FT_OSCURO
    c=ws_tip.cell(PIV_R+1, 2+i, tip)
    c.fill=fill(bg_tip); c.font=fnt(True,9,ft_tip)
    c.alignment=aln('center'); c.border=brd(C_BORDER_H)
hdr(ws_tip, PIV_R+1, 2+len(tipologias_orden), 'Total Unid.')
ws_tip.row_dimensions[PIV_R+1].height=22

piv = ru_df.pivot_table(index='Etiqueta', columns='Tipologia', values='Id', aggfunc='count', fill_value=0)
for r_i, etq in enumerate(etiquetas):
    if etq not in piv.index: continue
    row = PIV_R + 2 + r_i
    is_sub = etq.startswith('ST-')
    bg_lbl = 'D6E4F0' if is_sub else (C_ROW_ODD if r_i%2==0 else C_ROW_EVEN)
    c=ws_tip.cell(row,1,etq)
    c.fill=fill(bg_lbl); c.font=fnt(is_sub,10)
    c.alignment=aln('center'); c.border=brd()

    total_row = 0
    for i,tip in enumerate(tipologias_orden):
        n = int(piv.loc[etq, tip]) if tip in piv.columns else 0
        total_row += n
        c=ws_tip.cell(row, 2+i, n)
        c.fill=fill(bg_lbl); c.font=fnt(False,10)
        c.alignment=aln('center'); c.border=brd(); c.number_format='0'
    c=ws_tip.cell(row, 2+len(tipologias_orden), total_row)
    c.fill=fill(C_SUBTOT); c.font=fnt(True,10)
    c.alignment=aln('center'); c.border=brd(C_HDR_BG); c.number_format='0'
    ws_tip.row_dimensions[row].height=18

TOT_P = PIV_R + 2 + len(etiquetas)
tot(ws_tip, TOT_P, 1, 'TOTAL', ha='center')
gran_total = 0
for i,tip in enumerate(tipologias_orden):
    n = int(piv[tip].sum()) if tip in piv.columns else 0
    gran_total += n
    tot(ws_tip, TOT_P, 2+i, n, '0', 'center')
tot(ws_tip, TOT_P, 2+len(tipologias_orden), gran_total, '0', 'center')
ws_tip.row_dimensions[TOT_P].height=20
ws_tip.freeze_panes='B3'
# ── Guardar ───────────────────────────────────────────────
wb.save(output_path)
venta_v = df['SV'].sum()
const_v = df[df['Integra']]['GFA'].sum()
efic    = venta_v/const_v if const_v>0 else 0
semaf   = '🟢 Alta' if efic>0.75 else ('🟡 Media' if efic>0.5 else '🔴 Baja')
print(f"OK: {output_path}")
print(f"Venta: {venta_v:,.0f} m² | Construido: {const_v:,.0f} m² | Eficiencia: {efic:.1%} → {semaf}")
