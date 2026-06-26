"""
mobil-forma — motor de cabida (in-memory, apto Pyodide).

Adaptado de genera_cabida.py v1.8 (CLI) sin alterar la lógica de cálculo ni el
formato del Excel. Cambios respecto al original:
  - Recibe el CSV como texto y n_sub como argumento (sin sys.argv ni disco).
  - Devuelve el .xlsx como string base64 (cruza fácil el puente Python↔JS).
  - etiqueta_fn pasa a ser closure para ver n_sub.

Uso:
    b64 = generar_cabida(csv_text, n_sub)   # -> str base64 del xlsx
"""
import base64
import io
import re
import unicodedata

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colores canónicos por función ─────────────────────────
# Familias diferenciadas: Residencial (azules), Comercial (naranja/café),
# Oficinas (verdes escalonados), Otros usos (gris-azulado, lila, gris medio).
COLOR_CANONICO = {
    'Residencial Util':    '2E74B5',  # azul medio
    'Residencial Comun':   '1F3864',  # azul muy oscuro
    'Residencial Terraza': '5B9BD5',  # azul claro
    'Residencial Loggia':  'A9CCE3',  # azul pálido
    'Comercial Util':      'F4802A',  # naranja
    'Comercial Comun':     '843C0C',  # café
    'Oficinas Util':       '92D050',  # lima
    'Oficinas Comun':      '375623',  # verde muy oscuro
    'Estacionamientos':    '7F8FA6',  # gris azulado medio
    'Ascensores':          '9E7BB5',  # lila / violeta
    'Otro':                'A6A6A6',  # gris medio
}
COLOR_TINT = {
    'Residencial Util':    'D6E4F4',
    'Residencial Comun':   'D0D9E8',
    'Residencial Terraza': 'DCEBF7',
    'Residencial Loggia':  'EAF2FA',
    'Comercial Util':      'FDE8D8',
    'Comercial Comun':     'E8D5CD',
    'Oficinas Util':       'E9F5D9',
    'Oficinas Comun':      'D5E2CF',
    'Estacionamientos':    'E6E9EF',
    'Ascensores':          'EEE7F3',
    'Otro':                'ECECEC',
}
FT_BLANCO  = 'FFFFFF'
FT_OSCURO  = '1A1A1A'
# Funciones cuyo color de fondo es oscuro -> texto blanco en encabezados
HDR_OSCUROS = {'Residencial Util','Residencial Comun','Comercial Util','Comercial Comun',
               'Oficinas Comun','Estacionamientos','Ascensores'}

C_HDR_BG   = '1F3864'
C_HDR_FT   = 'FFFFFF'
C_TOTAL_BG = '2E74B5'
C_TOTAL_FT = 'FFFFFF'
C_SUBTOT   = 'D6E4F0'
C_ROW_ODD  = 'F7FAFD'
C_ROW_EVEN = 'FFFFFF'
C_BORDER   = 'BDD7EE'
C_BORDER_H = '4472C4'

CANONICO = {
    'ResidencialÚtil':'Residencial Util','ResidencialUtil':'Residencial Util',
    'ResidencialComún':'Residencial Comun','ResidencialComun':'Residencial Comun',
    'ResidencialTerraza':'Residencial Terraza','ResidencialLoggia':'Residencial Loggia',
    'ComercialÚtil':'Comercial Util','ComercialUtil':'Comercial Util',
    'ComercialComún':'Comercial Comun','ComercialComun':'Comercial Comun',
    'OficinasÚtil':'Oficinas Util','OficinasUtil':'Oficinas Util',
    'OficinasComún':'Oficinas Comun','OficinasComun':'Oficinas Comun',
    'Estacionamientos':'Estacionamientos','Parking':'Estacionamientos',
    'CirculaciónVertical':'Ascensores','CirculacionVertical':'Ascensores',
    'Circulación Vertical':'Ascensores','Ascensores':'Ascensores','Core':'Ascensores',
}
FACTOR_VENTA = {
    'Residencial Util':1.0,'Residencial Comun':0.0,'Residencial Terraza':0.5,
    'Residencial Loggia':1.0,'Comercial Util':1.0,'Comercial Comun':0.0,
    'Oficinas Util':1.0,'Oficinas Comun':0.0,'Estacionamientos':0.0,
    'Ascensores':0.0,'Otro':0.0,
}

# Mapeo GFA -> Tipologia (Residencial Util)
# Editable segun proyecto. Tolerancia +/- 1 m2 por redondeo en Forma.
TIPOLOGIA_RANGOS = [
    # (gfa_min, gfa_max, nombre_tipologia, dorm, banos, programa)
    (40, 50, '1D1B', 1, 1, '1 dorm + 1 bano'),
    (60, 80, '2D2B', 2, 2, '2 dorm + 2 banos'),
    (81, 95, '3D2B', 3, 2, '3 dorm + 2 banos'),
    (96, 120, '3D2B-XL', 3, 2, '3 dorm + 2 banos (XL)'),
]
PROGRAMA_TIPOLOGIA = {t: prog for _,_,t,_,_,prog in TIPOLOGIA_RANGOS}
COLOR_TIPOLOGIA = {
    '1D1B':    'B4C7E7',
    '2D2B':    '8FAADC',
    '3D2B':    '2E74B5',
    '3D2B-XL': '1F3864',
    'Otro':    'BFBFBF',
}

def tipologia(gfa, canonico):
    '''Devuelve la tipologia segun GFA. Solo aplica a Residencial Util.'''
    if canonico != 'Residencial Util': return ''
    g = float(gfa) if gfa else 0
    # Excluir unidades sin volumen (areas remanentes en Forma)
    if g <= 0: return ''
    for gmin, gmax, nombre, _, _, _ in TIPOLOGIA_RANGOS:
        if gmin <= g <= gmax: return nombre
    return f'Otro ({g:.0f}m2)'

CONSTRUIDO = {k:True for k in FACTOR_VENTA}
CONSTRUIDO['Ascensores'] = False
CONSTRUIDO['Otro']       = False

ORDEN_FUNC = [
    'Residencial Util','Residencial Comun','Residencial Terraza','Residencial Loggia',
    'Comercial Util','Comercial Comun','Oficinas Util','Oficinas Comun',
    'Estacionamientos','Ascensores','Otro'
]
GRUPOS = {
    'Residencial': ['Residencial Util','Residencial Comun','Residencial Terraza','Residencial Loggia'],
    'Comercial':   ['Comercial Util','Comercial Comun'],
    'Oficinas':    ['Oficinas Util','Oficinas Comun'],
    'Otros usos':  ['Estacionamientos','Ascensores','Otro'],
}

def canon(f):
    v = unicodedata.normalize('NFC', str(f).strip())
    if v in CANONICO: return CANONICO[v]
    # Intentar tambien sin espacios internos (Forma a veces exporta "Residencial Util" con espacio)
    v_compact = v.replace(' ', '')
    if v_compact in CANONICO: return CANONICO[v_compact]
    vs = ''.join(c for c in unicodedata.normalize('NFD',v) if unicodedata.category(c)!='Mn')
    vs_compact = vs.replace(' ', '')
    for k,c2 in CANONICO.items():
        ks = ''.join(ch for ch in unicodedata.normalize('NFD',k) if unicodedata.category(ch)!='Mn')
        if vs.lower()==ks.lower(): return c2
        if vs_compact.lower()==ks.lower(): return c2
    return 'Otro'

def extraer_nivel(s):
    m = re.search(r'/(\d+)_', str(s))
    if m: return int(m.group(1))
    return None

def extraer_edificio(s):
    """Edificio = fragmento del Id confinado entre el primer y el segundo '/'.

    Ej: 'root/54cce97/0_unit1' -> '54cce97'. Devuelve None si el Id no tiene esa
    forma (p. ej. filas manuales sin ruta)."""
    parts = str(s).split('/')
    if len(parts) >= 3 and parts[1] != '':
        return parts[1]
    return None

def sort_etq(e):
    """Orden de pisos: ST altos primero (más profundo), luego N1→N alto."""
    e = str(e)
    if e.startswith('ST-'):
        try: return -1000 + int(e[3:])
        except: return -999
    elif e.startswith('N'):
        try: return int(e[1:])
        except: return 9999
    return 9998

def sort_etq_tabla(e):
    """Orden para la TABLA: pisos altos arriba (N alto→N1), subterráneos abajo
    desde el menos profundo (ST-1, ST-2, …). N6,N5,…,N1, ST-1, ST-2."""
    e = str(e)
    if e.startswith('N'):
        try: return -int(e[1:])      # N6→-6 … N1→-1 (N alto primero)
        except: return -0.5
    elif e.startswith('ST-'):
        try: return 1000 + int(e[3:])  # ST-1→1001, ST-2→1002 (al fondo, en orden)
        except: return 1999
    return 500  # S/N entre medio

def es_departamento(tipo):
    """True si el elemento es una unidad habitacional (departamento), según la
    columna Tipo que exporta Forma. Robusto a acentos/mayúsculas/espacios."""
    t = ''.join(c for c in unicodedata.normalize('NFD', str(tipo))
                if unicodedata.category(c) != 'Mn').lower().strip()
    return ' '.join(t.split()) == 'unidad habitacional'

def limpiar_gfa(v):
    s = str(v).replace('\xa0','').replace(' ','').replace(' ','')
    return pd.to_numeric(s, errors='coerce')

# ── Lista de funciones canónicas válidas (para la UI de reclasificación) ──
FUNCIONES_CANONICAS = list(ORDEN_FUNC)


def _leer_csv(csv_text: str) -> pd.DataFrame:
    """Lee el CSV (texto) y normaliza headers ES/FR -> EN. Sin derivar columnas."""
    raw_text = csv_text.replace('\r\n', '\n').replace('\r', '\n')
    line = raw_text.split('\n')[0]
    sep = ';' if line.count(';') > line.count(',') else ','
    df = pd.read_csv(io.StringIO(raw_text), sep=sep)
    header_map = {
        'ID': 'Id', 'id': 'Id',
        'Tipo': 'Type', 'tipo': 'Type',
        'Función': 'Function', 'Funcion': 'Function', 'Fonction': 'Function',
        'función': 'Function', 'funcion': 'Function', 'fonction': 'Function',
    }
    df.rename(columns=header_map, inplace=True)
    return df


def _construir_df(csv_text: str, n_sub: int, ediciones=None):
    """(df, n_sub) base con columnas derivadas + ediciones del usuario aplicadas.

    `ediciones` (dict, opcional):
      {
        "n_sub": int,                      # opcional, sobrescribe n_sub
        "reclasificar": {Id: "Func"},      # cambia Function (raw) por una canónica
        "gfa":          {Id: float},        # sobrescribe GFA
        "nivel":        {Id: int},          # sobrescribe nivel_raw
        "excluir":      [Id, ...],          # filas a sacar del cálculo
        "agregar": [                        # filas manuales
           {"id": "manual-1", "funcion": "Residencial Terraza",
            "gfa": 12.5, "nivel": 6, "type": "Manual"}, ...
        ]
      }
    """
    n_sub = int(n_sub)
    ediciones = ediciones or {}
    if ediciones.get('n_sub') is not None:
        n_sub = int(ediciones['n_sub'])

    # Subterráneos por edificio (opcional). Mapa {id_edificio: n_sub_edificio}.
    # Si falta o un edificio no está en el mapa, se usa el n_sub global (compat).
    sub_por_ed = ediciones.get('subterraneos') or {}
    sub_por_ed = {str(k): int(v) for k, v in sub_por_ed.items()}

    def n_sub_de(ed):
        """Subterráneos del edificio dado; cae al global si no está mapeado."""
        if ed is not None and str(ed) in sub_por_ed:
            return sub_por_ed[str(ed)]
        return n_sub

    def etiqueta_fn(x, ed=None):
        if x is None or (isinstance(x, float) and pd.isna(x)): return 'S/N'
        x = int(x)
        ns = n_sub_de(ed)
        if x < ns: return f'ST-{ns - x}'
        return f'N{x - ns + 1}'

    df = _leer_csv(csv_text)
    if 'Id' not in df.columns:   df['Id'] = ['row-%d' % i for i in range(len(df))]
    if 'Type' not in df.columns: df['Type'] = ''
    if 'Function' not in df.columns: df['Function'] = ''
    df['Id'] = df['Id'].astype(str)

    df['GFA']       = df['GFA'].apply(limpiar_gfa).fillna(0)
    df['nivel_raw'] = df['Id'].apply(extraer_nivel)
    df['Edificio']  = df['Id'].apply(extraer_edificio)

    # ── Aplicar ediciones por-Id ──
    reclasif = ediciones.get('reclasificar') or {}
    gfa_ed   = ediciones.get('gfa') or {}
    nivel_ed = ediciones.get('nivel') or {}
    excluir  = set(ediciones.get('excluir') or [])

    if reclasif:
        df['Function'] = df.apply(
            lambda r: reclasif.get(r['Id'], r['Function']), axis=1)
    if gfa_ed:
        df['GFA'] = df.apply(
            lambda r: float(gfa_ed[r['Id']]) if r['Id'] in gfa_ed else r['GFA'], axis=1)
    if nivel_ed:
        df['nivel_raw'] = df.apply(
            lambda r: int(nivel_ed[r['Id']]) if r['Id'] in nivel_ed else r['nivel_raw'], axis=1)

    # ── Filas manuales agregadas ──
    agregar = ediciones.get('agregar') or []
    if agregar:
        nuevas = []
        for a in agregar:
            nv = a.get('nivel')
            nuevas.append({
                'Id': str(a.get('id', 'manual')),
                'Type': a.get('type', 'Manual'),
                'Function': a.get('funcion', 'Otro'),
                'GFA': float(a.get('gfa', 0) or 0),
                'nivel_raw': (int(nv) if nv is not None and str(nv) != '' else None),
                # Edificio destino (para subterráneos correctos y agrupación por edificio).
                'Edificio': (str(a['edificio']) if a.get('edificio') else None),
            })
        if nuevas:
            df = pd.concat([df, pd.DataFrame(nuevas)], ignore_index=True)

    # ── Excluir ──
    if excluir:
        df = df[~df['Id'].isin(excluir)].reset_index(drop=True)

    # ── Derivar el resto ──
    # La etiqueta de piso depende del edificio (subterráneos por edificio).
    df['Etiqueta'] = df.apply(
        lambda r: etiqueta_fn(r['nivel_raw'], r.get('Edificio')), axis=1)
    # Nombre legible del edificio (default "Edificio N" por orden de aparición).
    nombres_ed = {str(k): str(v) for k, v in (ediciones.get('nombres_edificio') or {}).items()}
    if nombres_ed:
        df['EdificioNombre'] = df['Edificio'].apply(
            lambda e: nombres_ed.get(str(e), e if e is not None else 'Sin edificio'))
    else:
        df['EdificioNombre'] = df['Edificio']
    df['Canonico'] = df['Function'].apply(canon)
    df['FV']       = df['Canonico'].map(FACTOR_VENTA).fillna(0)
    df['SV']       = df['GFA'] * df['FV']
    df['Integra']  = df['Canonico'].map(CONSTRUIDO).fillna(False)
    return df, n_sub


def tabla_elementos(csv_text: str, n_sub: int, ediciones=None) -> dict:
    """Lista de elementos para la tabla editable de la UI.

    Devuelve filas con Id, Type, Function (raw), Canonica, GFA, nivel, Etiqueta,
    si integra construido y si la función original no se reconoció ('Otro').
    """
    df, n_sub = _construir_df(csv_text, n_sub, ediciones)
    filas = []
    for row in df.itertuples(index=False):
        nivel = getattr(row, 'nivel_raw', None)
        filas.append({
            'id':        str(row.Id),
            'type':      '' if pd.isna(getattr(row, 'Type', '')) else str(getattr(row, 'Type', '')),
            'function':  '' if pd.isna(row.Function) else str(row.Function),
            'canonica':  str(row.Canonico),
            'gfa':       round(float(row.GFA), 2),
            'nivel':     (None if nivel is None or (isinstance(nivel, float) and pd.isna(nivel)) else int(nivel)),
            'etiqueta':  str(row.Etiqueta),
            'integra':   bool(row.Integra),
            'es_otro':   str(row.Canonico) == 'Otro',
            'es_manual': str(row.Id).startswith('manual'),
        })
    return {
        'filas': filas,
        'funciones_canonicas': FUNCIONES_CANONICAS,
        'n_otro': int((df['Canonico'] == 'Otro').sum()),
    }


def venta_por_funcion(csv_text: str, n_sub: int, ediciones=None) -> dict:
    """Superficie de venta (SV) agregada por función canónica.

    Devuelve solo las funciones con venta > 0 (las de factor 0 se omiten),
    ordenadas de mayor a menor, con su color y % sobre el total de venta.
    """
    df, n_sub = _construir_df(csv_text, n_sub, ediciones)
    sv = df.groupby('Canonico')['SV'].sum()
    total = float(sv.sum())
    items = []
    for fn in ORDEN_FUNC:
        v = float(sv.get(fn, 0.0))
        if v <= 0:
            continue
        items.append({
            'funcion': fn,
            'venta': round(v, 2),
            'pct': round(v / total, 4) if total > 0 else 0.0,
            'color': '#' + COLOR_CANONICO.get(fn, 'BFBFBF'),
        })
    items.sort(key=lambda x: x['venta'], reverse=True)
    return {'items': items, 'total': round(total, 2)}


def gfa_estacionamientos(csv_text: str, n_sub: int, ediciones=None) -> dict:
    """GFA de Estacionamientos, separado en superficie (sobre NTN) y subterráneo.

    Devuelve el total (compat) y el desglose por nivel (N-x = superficie, ST-x = subterráneo).
    """
    df, n_sub = _construir_df(csv_text, n_sub, ediciones)
    est = df[df['Canonico'] == 'Estacionamientos']
    sub_mask = est['Etiqueta'].astype(str).str.startswith('ST-')
    gfa_subt = float(est[sub_mask]['GFA'].sum())
    n_subt = int(sub_mask.sum())
    gfa_sup = float(est[~sub_mask]['GFA'].sum())
    n_sup = int((~sub_mask).sum())
    return {
        'gfa': round(gfa_sup + gfa_subt, 2),
        'n_elementos': int(len(est)),
        'superficie': {'gfa': round(gfa_sup, 2), 'n_elementos': n_sup},
        'subterraneo': {'gfa': round(gfa_subt, 2), 'n_elementos': n_subt},
    }


def inspeccionar_csv(csv_text: str) -> dict:
    """Inspección rápida del CSV SIN calcular cabida (para el asistente de carga).

    Devuelve los edificios en orden de aparición, con los niveles crudos detectados
    (los del Id, sin aplicar subterráneos) y el conteo de elementos. Alimenta la UI
    que pide nombrar edificios y fijar sus subterráneos antes de calcular.
    """
    df = _leer_csv(csv_text)
    if 'Id' not in df.columns:
        df['Id'] = ['row-%d' % i for i in range(len(df))]
    df['Id'] = df['Id'].astype(str)
    df['Edificio'] = df['Id'].apply(extraer_edificio)
    df['nivel_raw'] = df['Id'].apply(extraer_nivel)

    edificios = []
    vistos = []
    for ed in df['Edificio'].tolist():
        clave = 'Sin edificio' if ed is None else str(ed)
        if clave not in vistos:
            vistos.append(clave)
    for i, clave in enumerate(vistos, 1):
        if clave == 'Sin edificio':
            sub = df[df['Edificio'].isna()]
        else:
            sub = df[df['Edificio'] == clave]
        niveles = sorted({int(n) for n in sub['nivel_raw'].tolist()
                          if n is not None and not (isinstance(n, float) and pd.isna(n))})
        edificios.append({
            'id':            clave,
            'nombre_default': clave if clave == 'Sin edificio' else f'Edificio {i}',
            'pisos_raw':     niveles,
            'n_pisos':       len(niveles),
            'n_elementos':   int(len(sub)),
        })
    return {
        'n_edificios': len([e for e in edificios if e['id'] != 'Sin edificio']),
        'edificios':   edificios,
    }


def cabida_por_edificio(csv_text: str, n_sub: int, ediciones=None) -> dict:
    """Agrupa las unidades por edificio (fragmento del Id entre slashes).

    Por cada edificio devuelve el desglose por función canónica con:
      - unidades (nº de elementos), gfa (m²), venta (SV) y pct (% del GFA del
        edificio). Incluye el total por edificio y el total del proyecto.
    Los elementos sin edificio identificable se agrupan bajo 'Sin edificio'.
    """
    df, n_sub = _construir_df(csv_text, n_sub, ediciones)
    df = df.copy()
    df['Edificio'] = df['Edificio'].fillna('Sin edificio')
    df['_depto']   = df['Type'].apply(es_departamento)

    funciones_proyecto = [f for f in ORDEN_FUNC if f in df['Canonico'].unique()]

    # Orden de edificios por GFA total (mayor a menor), con 'Sin edificio' al final.
    gfa_por_ed = df.groupby('Edificio')['GFA'].sum().sort_values(ascending=False)
    orden_ed = [e for e in gfa_por_ed.index if e != 'Sin edificio']
    if 'Sin edificio' in gfa_por_ed.index:
        orden_ed.append('Sin edificio')

    # Mapa hash → nombre legible (lo trae _construir_df en EdificioNombre).
    nombre_de = {}
    for h, nm in zip(df['Edificio'], df['EdificioNombre']):
        nombre_de.setdefault(str(h), str(nm))

    edificios = []
    for ed in orden_ed:
        sub = df[df['Edificio'] == ed]
        total_gfa = float(sub['GFA'].sum())
        funcs = []
        for fn in funciones_proyecto:
            f_sub = sub[sub['Canonico'] == fn]
            if len(f_sub) == 0:
                continue
            g = float(f_sub['GFA'].sum())
            funcs.append({
                'funcion':  fn,
                'unidades': int(len(f_sub)),          # elementos de la función
                'deptos':   int(f_sub['_depto'].sum()),
                'gfa':      round(g, 2),
                'venta':    round(float(f_sub['SV'].sum()), 2),
                'pct':      round(g / total_gfa, 4) if total_gfa > 0 else 0.0,
                'color':    '#' + COLOR_CANONICO.get(fn, 'BFBFBF'),
            })
        construido = float(sub[sub['Integra']]['GFA'].sum())
        venta = float(sub['SV'].sum())
        edificios.append({
            'edificio':     nombre_de.get(str(ed), str(ed)),
            'edificio_id':  str(ed),
            'departamentos': int(sub['_depto'].sum()),
            'unidades':     int(len(sub)),            # total de elementos
            'gfa':          round(total_gfa, 2),
            'construido':   round(construido, 2),
            'venta':        round(venta, 2),
            'eficiencia':   round(venta / construido, 4) if construido > 0 else 0.0,
            'funciones':    funcs,
        })

    tot_constr = float(df[df['Integra']]['GFA'].sum())
    tot_venta = float(df['SV'].sum())
    return {
        'n_edificios': len(orden_ed),
        'funciones':   funciones_proyecto,
        'edificios':   edificios,
        'total': {
            'departamentos': int(df['_depto'].sum()),
            'unidades': int(len(df)),
            'gfa':      round(float(df['GFA'].sum()), 2),
            'construido': round(tot_constr, 2),
            'venta':    round(tot_venta, 2),
            'eficiencia': round(tot_venta / tot_constr, 4) if tot_constr > 0 else 0.0,
        },
    }


def paleta_canonica() -> list:
    """Paleta de funciones canónicas con su color (hex y rgb)."""
    out = []
    for fn in ORDEN_FUNC:
        h = COLOR_CANONICO.get(fn, 'BFBFBF')
        r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
        out.append({
            'funcion': fn,
            'hex': '#' + h,
            'rgb': [r, g, b],
            'rgb_str': f'rgb({r}, {g}, {b})',
        })
    return out


def paleta_xlsx() -> str:
    """Excel «tipo Mobil» de la paleta canónica: Función · Color · RGB.

    Cada fila lleva una celda rellena con el color real de la función. Devuelve
    el .xlsx codificado en base64 (mismo puente que generar_cabida).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Paleta'
    ws.sheet_view.showGridLines = False

    # Título (banda Mobil)
    ws.merge_cells('A1:C1')
    c = ws.cell(1, 1, 'PALETA DE FUNCIONES CANÓNICAS')
    c.fill = fill(C_HDR_BG); c.font = fnt(True, 13, C_HDR_FT); c.alignment = aln('center')
    ws.row_dimensions[1].height = 28

    # Encabezados
    for i, h in enumerate(['Función canónica', 'Color', 'RGB'], 1):
        hdr(ws, 2, i, h, ha='left' if i != 2 else 'center')
    ws.row_dimensions[2].height = 20

    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 20

    # Filas
    for idx, p in enumerate(paleta_canonica(), 3):
        h = p['hex'].lstrip('#')
        r, g, b = p['rgb']
        bg = C_ROW_ODD if idx % 2 == 0 else C_ROW_EVEN

        ca = ws.cell(idx, 1, p['funcion'])
        ca.fill = fill(bg); ca.font = fnt(False, 10, '1A1A1A')
        ca.alignment = aln('left'); ca.border = brd()

        # Celda de color: relleno = color de la función
        cb = ws.cell(idx, 2, '')
        cb.fill = fill(h); cb.border = brd()

        cc = ws.cell(idx, 3, f'{r}, {g}, {b}')
        cc.fill = fill(bg); cc.font = fnt(False, 10, '555555', name='Consolas')
        cc.alignment = aln('left'); cc.border = brd()
        ws.row_dimensions[idx].height = 22

    buf = io.BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode('ascii')


def fill(c):  return PatternFill('solid', fgColor=c)
def fnt(bold=False, sz=10, color='1A1A1A', name='Arial Narrow'):
    return Font(name=name, bold=bold, size=sz, color=color)
def brd(color=C_BORDER):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def aln(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def hdr(ws, r, c, val, bg=C_HDR_BG, ft=C_HDR_FT, sz=10, bold=True, wrap=False, ha='center'):
    cell = ws.cell(r, c, val)
    cell.fill=fill(bg); cell.font=fnt(bold,sz,ft)
    cell.alignment=aln(ha,wrap=wrap); cell.border=brd(C_BORDER_H)
    return cell

def tot(ws, r, c, val, fmt=None, ha='right'):
    cell = ws.cell(r, c, val)
    cell.fill=fill(C_TOTAL_BG); cell.font=fnt(True,10,C_TOTAL_FT)
    cell.alignment=aln(ha); cell.border=brd(C_HDR_BG)
    if fmt: cell.number_format=fmt
    return cell


def generar_cabida(csv_text: str, n_sub: int, ediciones=None) -> str:
    """Genera el Excel de cabida a partir del CSV (texto), n_sub y ediciones.

    Devuelve el contenido del .xlsx codificado en base64.
    """
    df, n_sub = _construir_df(csv_text, n_sub, ediciones)

    # === Tipologias (Residencial Util) y reparto de terraza ===
    df['building'] = df['Id'].astype(str).str.extract(r'root/([^/]+)/')
    df['Tipologia'] = df.apply(lambda r: tipologia(r['GFA'], r['Canonico']), axis=1)

    ru_mask = df['Canonico']=='Residencial Util'
    ter_mask = df['Canonico']=='Residencial Terraza'

    terraza_total_proyecto = df[ter_mask]['GFA'].sum()
    gfa_util_total_proyecto = df[ru_mask]['GFA'].sum()
    factor_terraza = terraza_total_proyecto / gfa_util_total_proyecto if gfa_util_total_proyecto > 0 else 0

    def terraza_unidad(row):
        if row['Canonico']!='Residencial Util': return 0.0
        return round(row['GFA'] * factor_terraza, 2)

    df['Terraza_Unidad'] = df.apply(terraza_unidad, axis=1)
    df['GFA_Total_Unidad'] = df['GFA'] + df['Terraza_Unidad']
    df['SV_Unidad'] = df['GFA'] + 0.5 * df['Terraza_Unidad']

    funciones_presentes = [f for f in ORDEN_FUNC if f in df['Canonico'].unique()]

    def sort_etq(e):
        if e.startswith('ST-'):
            try: return -1000 + int(e[3:])
            except: return -999
        elif e.startswith('N'):
            try: return int(e[1:])
            except: return 9999
        return 9998

    etiquetas = sorted(df['Etiqueta'].unique().tolist(), key=sort_etq)
    df_sorted = df.sort_values(['nivel_raw','Canonico']).reset_index(drop=True)
    n_data    = len(df_sorted)

    wb = Workbook()

    # Forzar cálculo automático + recálculo completo al abrir.
    # El libro guarda solo fórmulas (openpyxl no evalúa), así que depende de que
    # Excel recalcule al abrirlo. Si el Excel del usuario está en modo Manual
    # (típico tras un cierre forzado / recuperación de documentos), las celdas con
    # fórmula se ven en blanco. Fijar 'auto' aquí lo evita en cualquier equipo.
    wb.calculation.calcMode = 'auto'
    wb.calculation.fullCalcOnLoad = True

    # ═══════════════════════════════════════════════════════
    # HOJA 1 — DATOS
    # ═══════════════════════════════════════════════════════
    ws_d = wb.active
    ws_d.title = 'Datos'

    ws_d.merge_cells('A1:J1')
    c=ws_d.cell(1,1,'DATOS — DETALLE POR ELEMENTO  |  Fuente de Cabida y Tipología')
    c.fill=fill(C_HDR_BG); c.font=fnt(True,13,C_HDR_FT); c.alignment=aln('center')
    ws_d.row_dimensions[1].height=28

    DATOS_HDRS  = ['Id','Type','Function (raw)','GFA m²','Nivel','Etiqueta',
                   'Función Canónica','Factor Venta','Sup. Venta m²','Integra Construido']
    DATOS_WIDTHS= [42,14,24,12,8,10,22,14,16,18]
    for i,(h,w) in enumerate(zip(DATOS_HDRS,DATOS_WIDTHS),1):
        hdr(ws_d,2,i,h); ws_d.column_dimensions[get_column_letter(i)].width=w
    ws_d.row_dimensions[2].height=20

    for r_i, row in enumerate(df_sorted.itertuples(index=False), 3):
        is_sub = str(row.Etiqueta).startswith('ST-')
        bg = 'D6E4F0' if is_sub else (C_ROW_ODD if r_i%2==0 else C_ROW_EVEN)
        bld = is_sub
        vals=[
            (row.Id,        '@',          'left'),
            (getattr(row,'Type',''), '@', 'left'),
            (row.Function,  '@',          'left'),
            (row.GFA,       '#,##0.00',   'right'),
            (row.nivel_raw, '0',          'center'),
            (row.Etiqueta,  '@',          'center'),
            (row.Canonico,  '@',          'left'),
            (row.FV,        '0%',         'center'),
            (row.SV,        '#,##0.00',   'right'),
            ('✓' if row.Integra else '—','@','center'),
        ]
        for c_i,(val,fmt,ha) in enumerate(vals,1):
            cell=ws_d.cell(r_i,c_i,val)
            cell.fill=fill(bg); cell.font=fnt(bld,9)
            cell.alignment=aln(ha); cell.border=brd(); cell.number_format=fmt
        ws_d.row_dimensions[r_i].height=16

    TR = 3 + n_data   # fila total Datos
    tot(ws_d,TR,1,'TOTAL',ha='center')
    for ci in range(2,11): tot(ws_d,TR,ci,None)
    tot(ws_d,TR,4, df_sorted['GFA'].sum(),'#,##0.00')
    tot(ws_d,TR,9, df_sorted['SV'].sum(), '#,##0.00')
    ws_d.row_dimensions[TR].height=20
    ws_d.freeze_panes='A3'

    # ── Agregados precalculados ──────────────────────────────
    # Cabida y Programa se escriben como VALORES (no fórmulas) para que NUNCA
    # dependan del modo de cálculo de Excel (en modo Manual las fórmulas salían
    # en blanco). Estos sums replican exactamente lo que hacían los SUMIFS.
    gfa_ec  = df.groupby(['Etiqueta', 'Canonico'])['GFA'].sum()
    gfa_fn  = df.groupby('Canonico')['GFA'].sum()
    sv_fn   = df.groupby('Canonico')['SV'].sum()
    sv_etq  = df.groupby('Etiqueta')['SV'].sum()
    construido_fns = [fn for fn in funciones_presentes if CONSTRUIDO.get(fn, False)]
    tot_construido_gfa = float(sum(float(gfa_fn.get(fn, 0.0)) for fn in construido_fns))
    sv_total  = float(df['SV'].sum())
    gfa_total = float(df['GFA'].sum())
    def _gfa_ec(etq, fn):
        try: return float(gfa_ec.loc[(etq, fn)])
        except KeyError: return 0.0

    # ═══════════════════════════════════════════════════════
    # HOJA 2 — CABIDA
    # ═══════════════════════════════════════════════════════
    ws_c = wb.create_sheet('Cabida')

    fn_cols = {}
    col_ptr = 2
    for fn in funciones_presentes:
        fn_cols[fn] = col_ptr; col_ptr += 1
    COL_TOTAL = col_ptr; col_ptr += 1
    COL_VENTA = col_ptr; col_ptr += 1
    LAST_COL  = col_ptr - 1

    ws_c.merge_cells(f'A1:{get_column_letter(LAST_COL)}1')
    c=ws_c.cell(1,1,'CABIDA NUMÉRICA — SUPERFICIES POR NIVEL')
    c.fill=fill(C_HDR_BG); c.font=fnt(True,13,C_HDR_FT)
    c.alignment=aln('center'); ws_c.row_dimensions[1].height=28

    # Fila 2 — grupos
    col_g = 2
    for grp, fns in GRUPOS.items():
        fns_p = [f for f in fns if f in funciones_presentes]
        if not fns_p: continue
        c1=col_g; c2=col_g+len(fns_p)-1
        if c2>c1: ws_c.merge_cells(f'{get_column_letter(c1)}2:{get_column_letter(c2)}2')
        cell=ws_c.cell(2,c1,grp)
        cell.fill=fill(C_TOTAL_BG); cell.font=fnt(True,9,C_HDR_FT)
        cell.alignment=aln('center'); cell.border=brd(C_BORDER_H)
        col_g += len(fns_p)
    for lbl,col in [('TOTAL m²',COL_TOTAL),('VENTA m²',COL_VENTA)]:
        cell=ws_c.cell(2,col,lbl)
        cell.fill=fill(C_TOTAL_BG); cell.font=fnt(True,9,C_HDR_FT)
        cell.alignment=aln('center'); cell.border=brd(C_BORDER_H)

    # Fila 3 — "Nivel" + nombres función coloreados
    ws_c.merge_cells('A2:A3')
    c=ws_c.cell(2,1,'Nivel')
    c.fill=fill(C_HDR_BG); c.font=fnt(True,10,C_HDR_FT)
    c.alignment=aln('center','center'); c.border=brd(C_BORDER_H)
    ws_c.column_dimensions['A'].width=10

    for fn, col in fn_cols.items():
        bg  = COLOR_CANONICO.get(fn,'BFBFBF')
        ft  = FT_BLANCO if fn in HDR_OSCUROS else FT_OSCURO
        cell=ws_c.cell(3,col,fn)
        cell.fill=fill(bg); cell.font=fnt(True,8,ft)
        cell.alignment=aln('center',wrap=True); cell.border=brd(C_BORDER_H)
        ws_c.column_dimensions[get_column_letter(col)].width=15

    for col,lbl in [(COL_TOTAL,'Total m²'),(COL_VENTA,'Venta m²')]:
        cell=ws_c.cell(3,col,lbl)
        cell.fill=fill(C_TOTAL_BG); cell.font=fnt(True,9,C_HDR_FT)
        cell.alignment=aln('center'); cell.border=brd(C_BORDER_H)
        ws_c.column_dimensions[get_column_letter(col)].width=13
    ws_c.row_dimensions[2].height=22; ws_c.row_dimensions[3].height=42

    # Filas de datos — SUMIFS vinculados a Datos
    DS = 4   # DATA_START en Cabida
    for r_i, etq in enumerate(etiquetas):
        row = DS + r_i
        is_sub = etq.startswith('ST-')
        bg_lbl = 'D6E4F0' if is_sub else (C_ROW_ODD if r_i%2==0 else C_ROW_EVEN)
        cell=ws_c.cell(row,1,etq)
        cell.fill=fill(bg_lbl); cell.font=fnt(is_sub,10)
        cell.alignment=aln('center'); cell.border=brd()

        for fn, col in fn_cols.items():
            tint = COLOR_TINT.get(fn, C_ROW_ODD)
            cell2=ws_c.cell(row,col,_gfa_ec(etq,fn))
            cell2.fill=fill(tint); cell2.font=fnt(False,10)
            cell2.alignment=aln('right'); cell2.border=brd(); cell2.number_format='#,##0'

        tot_etq = float(sum(_gfa_ec(etq,fn) for fn in construido_fns))
        c_tot=ws_c.cell(row,COL_TOTAL,tot_etq)
        c_tot.fill=fill(C_SUBTOT); c_tot.font=fnt(True,10)
        c_tot.alignment=aln('right'); c_tot.border=brd(C_HDR_BG); c_tot.number_format='#,##0'

        c_sv=ws_c.cell(row,COL_VENTA,float(sv_etq.get(etq,0.0)))
        c_sv.fill=fill(C_SUBTOT); c_sv.font=fnt(True,10)
        c_sv.alignment=aln('right'); c_sv.border=brd(C_HDR_BG); c_sv.number_format='#,##0'
        ws_c.row_dimensions[row].height=18

    # Fila TOTAL
    TOT_C = DS + len(etiquetas)
    tot(ws_c,TOT_C,1,'TOTAL',ha='center')
    for fn, col in fn_cols.items():
        c=ws_c.cell(TOT_C,col,float(gfa_fn.get(fn,0.0)))
        c.fill=fill(C_TOTAL_BG); c.font=fnt(True,10,C_TOTAL_FT)
        c.alignment=aln('right'); c.border=brd(C_HDR_BG); c.number_format='#,##0'
    for col,val in [(COL_TOTAL,tot_construido_gfa),(COL_VENTA,sv_total)]:
        c=ws_c.cell(TOT_C,col,float(val))
        c.fill=fill(C_TOTAL_BG); c.font=fnt(True,10,C_TOTAL_FT)
        c.alignment=aln('right'); c.border=brd(C_HDR_BG); c.number_format='#,##0'
    ws_c.row_dimensions[TOT_C].height=20

    # KPIs
    KR = TOT_C + 2
    eficiencia = (sv_total / tot_construido_gfa) if tot_construido_gfa > 0 else 0.0
    calificacion = ('🟢 Alta' if eficiencia > 0.75
                    else '🟡 Media' if eficiencia > 0.5 else '🔴 Baja')
    for r,(lbl,val,fmt) in enumerate([
        ('Sup. Venta',      sv_total,           '#,##0" m²"'),
        ('Sup. Construida', tot_construido_gfa, '#,##0" m²"'),
        ('Eficiencia',      eficiencia,         '0.0%'),
        ('Calificación',    calificacion,       '@'),
    ], KR):
        c1=ws_c.cell(r,1,lbl)
        c1.fill=fill('1F4E79'); c1.font=fnt(True,10,C_HDR_FT)
        c1.alignment=aln('left'); c1.border=brd(C_BORDER_H)
        c2=ws_c.cell(r,2,val)
        c2.fill=fill(C_SUBTOT); c2.font=fnt(True,10)
        c2.alignment=aln('right'); c2.border=brd(); c2.number_format=fmt

    ws_c.freeze_panes='B4'

    # ═══════════════════════════════════════════════════════
    # HOJA 3 — PROGRAMA
    # ═══════════════════════════════════════════════════════
    ws_t = wb.create_sheet('Programa')

    ws_t.merge_cells('A1:G1')
    c=ws_t.cell(1,1,'PROGRAMA — RESUMEN POR FUNCIÓN CANÓNICA')
    c.fill=fill(C_HDR_BG); c.font=fnt(True,13,C_HDR_FT); c.alignment=aln('center')
    ws_t.row_dimensions[1].height=28

    T_HDRS  = ['Función Canónica','GFA Total m²','Sup. Venta m²','Factor Venta','% Construido','Integra Construido','Color']
    T_WIDTHS= [24,16,16,14,14,18,12]
    for i,(h,w) in enumerate(zip(T_HDRS,T_WIDTHS),1):
        hdr(ws_t,2,i,h); ws_t.column_dimensions[get_column_letter(i)].width=w
    ws_t.row_dimensions[2].height=20

    fn_row_map = {fn: 3+i for i,fn in enumerate(funciones_presentes)}

    for fn in funciones_presentes:
        r   = fn_row_map[fn]
        odd = r % 2 == 0
        bg_fn = COLOR_CANONICO.get(fn,'BFBFBF')
        ft_fn = FT_BLANCO if fn in HDR_OSCUROS else FT_OSCURO
        integra = CONSTRUIDO.get(fn,False)
        fv      = FACTOR_VENTA.get(fn,0)
        bg_d    = C_ROW_ODD if odd else C_ROW_EVEN

        c=ws_t.cell(r,1,fn)
        c.fill=fill(bg_fn); c.font=fnt(True,10,ft_fn)
        c.alignment=aln('left'); c.border=brd(C_BORDER_H)

        fn_gfa = float(gfa_fn.get(fn,0.0))
        c=ws_t.cell(r,2,fn_gfa)
        c.fill=fill(bg_d); c.font=fnt(False,10)
        c.alignment=aln('right'); c.border=brd(); c.number_format='#,##0'

        c=ws_t.cell(r,3,float(sv_fn.get(fn,0.0)))
        c.fill=fill(bg_d); c.font=fnt(False,10)
        c.alignment=aln('right'); c.border=brd(); c.number_format='#,##0'

        c=ws_t.cell(r,4,fv)
        c.fill=fill(bg_d); c.font=fnt(False,10)
        c.alignment=aln('center'); c.border=brd(); c.number_format='0%'

        if integra:
            c=ws_t.cell(r,5,(fn_gfa/tot_construido_gfa) if tot_construido_gfa>0 else 0.0)
            c.number_format='0.0%'
        else:
            c=ws_t.cell(r,5,'—')
            c.number_format='@'
        c.fill=fill(bg_d); c.font=fnt(False,10)
        c.alignment=aln('center'); c.border=brd()

        c=ws_t.cell(r,6,'✓ Sí' if integra else '✗ No')
        c.fill=fill(bg_d); c.font=fnt(False,10)
        c.alignment=aln('center'); c.border=brd()

        c=ws_t.cell(r,7,f'#{bg_fn}')
        c.fill=fill(bg_fn); c.font=fnt(False,9,ft_fn)
        c.alignment=aln('center'); c.border=brd()
        ws_t.row_dimensions[r].height=18

    TOT_T = 3 + len(funciones_presentes)
    tot(ws_t,TOT_T,1,'TOTAL',ha='center')
    c=ws_t.cell(TOT_T,2,gfa_total)
    c.fill=fill(C_TOTAL_BG); c.font=fnt(True,10,C_TOTAL_FT)
    c.alignment=aln('right'); c.border=brd(C_HDR_BG); c.number_format='#,##0'
    c=ws_t.cell(TOT_T,3,sv_total)
    c.fill=fill(C_TOTAL_BG); c.font=fnt(True,10,C_TOTAL_FT)
    c.alignment=aln('right'); c.border=brd(C_HDR_BG); c.number_format='#,##0'
    for col in [4,5,6,7]:
        c=ws_t.cell(TOT_T,col,None)
        c.fill=fill(C_TOTAL_BG); c.font=fnt(True,10,C_TOTAL_FT)
        c.alignment=aln('center'); c.border=brd(C_HDR_BG)
    ws_t.row_dimensions[TOT_T].height=20
    ws_t.freeze_panes='A3'

    # ═══════════════════════════════════════════════════════
    # HOJA 4 — TIPOLOGIAS
    # ═══════════════════════════════════════════════════════
    ru_df = df[(df['Canonico']=='Residencial Util') & (df['GFA']>0)].copy()
    tipologias_presentes = [t for _,_,t,_,_,_ in TIPOLOGIA_RANGOS if t in ru_df['Tipologia'].unique()]
    otros_tipos = sorted([t for t in ru_df['Tipologia'].unique() if t not in [x[2] for x in TIPOLOGIA_RANGOS]])
    tipologias_orden = tipologias_presentes + otros_tipos

    ws_tip = wb.create_sheet('Tipologias')
    ws_tip.merge_cells('A1:I1')
    c=ws_tip.cell(1,1,'TIPOLOGIAS RESIDENCIALES — DETALLE POR UNIDAD Y NIVEL')
    c.fill=fill(C_HDR_BG); c.font=fnt(True,13,C_HDR_FT); c.alignment=aln('center')
    ws_tip.row_dimensions[1].height=28

    T_HDRS = ['Tipologia','Programa','N Unidades','GFA Util Prom m2','Terraza Prom m2','Total Prom m2','Sup Venta Prom m2','GFA Util Total','Sup Venta Total']
    T_W    = [12,22,12,16,16,16,16,16,16]
    for i,(h,w) in enumerate(zip(T_HDRS,T_W),1):
        hdr(ws_tip,2,i,h); ws_tip.column_dimensions[get_column_letter(i)].width=w
    ws_tip.row_dimensions[2].height=22

    for r_i, tip in enumerate(tipologias_orden):
        r = 3 + r_i
        sub = ru_df[ru_df['Tipologia']==tip]
        n_un = len(sub)
        gfa_prom = sub['GFA'].mean() if n_un>0 else 0
        ter_prom = sub['Terraza_Unidad'].mean() if n_un>0 else 0
        tot_prom = sub['GFA_Total_Unidad'].mean() if n_un>0 else 0
        sv_prom  = sub['SV_Unidad'].mean() if n_un>0 else 0
        gfa_tot  = sub['GFA'].sum()
        sv_tot   = sub['SV_Unidad'].sum()
        bg_tip   = COLOR_TIPOLOGIA.get(tip,'BFBFBF')
        is_dark  = tip in ('3D2B','3D2B-XL')
        ft_tip   = FT_BLANCO if is_dark else FT_OSCURO

        c=ws_tip.cell(r,1,tip)
        c.fill=fill(bg_tip); c.font=fnt(True,10,ft_tip)
        c.alignment=aln('center'); c.border=brd(C_BORDER_H)

        odd = r % 2 == 0
        bg_d = C_ROW_ODD if odd else C_ROW_EVEN
        programa = PROGRAMA_TIPOLOGIA.get(tip, '-')

        c=ws_tip.cell(r,2,programa)
        c.fill=fill(bg_d); c.font=fnt(False,10)
        c.alignment=aln('left'); c.border=brd()

        vals = [
            (n_un,    '0',         'center'),
            (gfa_prom,'#,##0.0',   'right'),
            (ter_prom,'#,##0.0',   'right'),
            (tot_prom,'#,##0.0',   'right'),
            (sv_prom, '#,##0.0',   'right'),
            (gfa_tot, '#,##0',     'right'),
            (sv_tot,  '#,##0',     'right'),
        ]
        for c_i,(v,fmt_,ha) in enumerate(vals,3):
            c=ws_tip.cell(r,c_i,v)
            c.fill=fill(bg_d); c.font=fnt(False,10)
            c.alignment=aln(ha); c.border=brd(); c.number_format=fmt_
        ws_tip.row_dimensions[r].height=18

    TOT_R = 3 + len(tipologias_orden)
    tot(ws_tip,TOT_R,1,'TOTAL',ha='center')
    tot_n   = len(ru_df)
    tot_gfa = ru_df['GFA'].sum()
    tot_sv  = ru_df['SV_Unidad'].sum()
    tot(ws_tip,TOT_R,2,None)
    tot(ws_tip,TOT_R,3,tot_n,'0','center')
    for ci in [4,5,6,7]: tot(ws_tip,TOT_R,ci,None)
    tot(ws_tip,TOT_R,8,tot_gfa,'#,##0','right')
    tot(ws_tip,TOT_R,9,tot_sv,'#,##0','right')
    ws_tip.row_dimensions[TOT_R].height=20

    # === Pivot Nivel x Tipologia ===
    PIV_R = TOT_R + 3
    ws_tip.merge_cells(start_row=PIV_R, start_column=1, end_row=PIV_R, end_column=len(tipologias_orden)+2)
    c=ws_tip.cell(PIV_R,1,'CONTEO DE UNIDADES POR TIPOLOGIA Y NIVEL')
    c.fill=fill(C_TOTAL_BG); c.font=fnt(True,11,C_TOTAL_FT); c.alignment=aln('center')
    ws_tip.row_dimensions[PIV_R].height=22

    hdr(ws_tip, PIV_R+1, 1, 'Nivel')
    for i,tip in enumerate(tipologias_orden):
        bg_tip = COLOR_TIPOLOGIA.get(tip,'BFBFBF')
        is_dark = tip in ('3D2B','3D2B-XL')
        ft_tip = FT_BLANCO if is_dark else FT_OSCURO
        c=ws_tip.cell(PIV_R+1, 2+i, tip)
        c.fill=fill(bg_tip); c.font=fnt(True,9,ft_tip)
        c.alignment=aln('center'); c.border=brd(C_BORDER_H)
    hdr(ws_tip, PIV_R+1, 2+len(tipologias_orden), 'Total Unid.')
    ws_tip.row_dimensions[PIV_R+1].height=22

    piv = ru_df.pivot_table(index='Etiqueta', columns='Tipologia', values='Id', aggfunc='count', fill_value=0)
    for r_i, etq in enumerate(etiquetas):
        if etq not in piv.index: continue
        row = PIV_R + 2 + r_i
        is_sub = etq.startswith('ST-')
        bg_lbl = 'D6E4F0' if is_sub else (C_ROW_ODD if r_i%2==0 else C_ROW_EVEN)
        c=ws_tip.cell(row,1,etq)
        c.fill=fill(bg_lbl); c.font=fnt(is_sub,10)
        c.alignment=aln('center'); c.border=brd()

        total_row = 0
        for i,tip in enumerate(tipologias_orden):
            n = int(piv.loc[etq, tip]) if tip in piv.columns else 0
            total_row += n
            c=ws_tip.cell(row, 2+i, n)
            c.fill=fill(bg_lbl); c.font=fnt(False,10)
            c.alignment=aln('center'); c.border=brd(); c.number_format='0'
        c=ws_tip.cell(row, 2+len(tipologias_orden), total_row)
        c.fill=fill(C_SUBTOT); c.font=fnt(True,10)
        c.alignment=aln('center'); c.border=brd(C_HDR_BG); c.number_format='0'
        ws_tip.row_dimensions[row].height=18

    TOT_P = PIV_R + 2 + len(etiquetas)
    tot(ws_tip, TOT_P, 1, 'TOTAL', ha='center')
    gran_total = 0
    for i,tip in enumerate(tipologias_orden):
        n = int(piv[tip].sum()) if tip in piv.columns else 0
        gran_total += n
        tot(ws_tip, TOT_P, 2+i, n, '0', 'center')
    tot(ws_tip, TOT_P, 2+len(tipologias_orden), gran_total, '0', 'center')
    ws_tip.row_dimensions[TOT_P].height=20
    ws_tip.freeze_panes='B3'

    # ═══════════════════════════════════════════════════════
    # HOJA 5 — POR EDIFICIO
    # ═══════════════════════════════════════════════════════
    ws_e = wb.create_sheet('Por Edificio')

    EDI_HDRS   = ['Función','Elem.','m² (GFA)','Venta m²','% del edif.']
    EDI_WIDTHS = [26,11,14,14,12]
    LAST_E = len(EDI_HDRS)

    ws_e.merge_cells(f'A1:{get_column_letter(LAST_E)}1')
    c=ws_e.cell(1,1,'CABIDA POR EDIFICIO — UNIDADES Y SUPERFICIE POR FUNCIÓN')
    c.fill=fill(C_HDR_BG); c.font=fnt(True,13,C_HDR_FT); c.alignment=aln('center')
    ws_e.row_dimensions[1].height=28
    for i,w in enumerate(EDI_WIDTHS,1):
        ws_e.column_dimensions[get_column_letter(i)].width=w

    # Agrupar por edificio (los sin edificio van al final)
    df_e = df.assign(Edificio=df['Edificio'].fillna('Sin edificio'),
                     _depto=df['Type'].apply(es_departamento))
    # Mapa hash → nombre legible (default = el propio hash si no se renombró).
    nombre_de = {}
    if 'EdificioNombre' in df.columns:
        for h, nm in zip(df_e['Edificio'], df['EdificioNombre'].fillna('Sin edificio')):
            nombre_de.setdefault(str(h), str(nm))
    gfa_por_ed = df_e.groupby('Edificio')['GFA'].sum().sort_values(ascending=False)
    orden_ed = [e for e in gfa_por_ed.index if e != 'Sin edificio']
    if 'Sin edificio' in gfa_por_ed.index:
        orden_ed.append('Sin edificio')

    r = 2
    for ed in orden_ed:
        sub = df_e[df_e['Edificio'] == ed]
        tot_gfa = float(sub['GFA'].sum())
        tot_uni = int(len(sub))
        tot_dep = int(sub['_depto'].sum())
        tot_sv  = float(sub['SV'].sum())
        ed_nombre = nombre_de.get(str(ed), str(ed))

        # Banda con el nombre del edificio + resumen
        ws_e.merge_cells(f'A{r}:{get_column_letter(LAST_E)}{r}')
        cell=ws_e.cell(r,1,f'{ed_nombre}    ·    {tot_dep} departamentos    ·    {tot_gfa:,.0f} m²')
        cell.fill=fill(C_HDR_BG); cell.font=fnt(True,11,C_HDR_FT); cell.alignment=aln('left')
        ws_e.row_dimensions[r].height=22
        r+=1

        # Encabezado de columnas del bloque
        for i,h in enumerate(EDI_HDRS,1):
            hdr(ws_e, r, i, h)
        ws_e.row_dimensions[r].height=18
        r+=1

        # Filas por función presente en el edificio
        for fn in funciones_presentes:
            f_sub = sub[sub['Canonico'] == fn]
            if len(f_sub) == 0: continue
            g=float(f_sub['GFA'].sum()); v=float(f_sub['SV'].sum()); u=int(len(f_sub))
            pct = g/tot_gfa if tot_gfa>0 else 0.0
            bg  = COLOR_CANONICO.get(fn,'BFBFBF')
            ft  = FT_BLANCO if fn in HDR_OSCUROS else FT_OSCURO
            tint= COLOR_TINT.get(fn,'ECECEC')
            celdas=[
                (fn,  '@',         'left',   bg,   ft,        True),
                (u,   '0',         'center', tint, FT_OSCURO, False),
                (g,   '#,##0.00',  'right',  tint, FT_OSCURO, False),
                (v,   '#,##0.00',  'right',  tint, FT_OSCURO, False),
                (pct, '0.0%',      'center', tint, FT_OSCURO, False),
            ]
            for ci,(val,fmt,ha,bgc,ftc,bold) in enumerate(celdas,1):
                cc=ws_e.cell(r,ci,val)
                cc.fill=fill(bgc); cc.font=fnt(bold,10,ftc)
                cc.alignment=aln(ha); cc.border=brd(); cc.number_format=fmt
            ws_e.row_dimensions[r].height=16
            r+=1

        # Subtotal del edificio
        tot(ws_e,r,1,'Subtotal',ha='left')
        tot(ws_e,r,2,tot_uni,'0','center')
        tot(ws_e,r,3,tot_gfa,'#,##0.00')
        tot(ws_e,r,4,tot_sv,'#,##0.00')
        tot(ws_e,r,5,1.0,'0.0%','center')
        ws_e.row_dimensions[r].height=18
        r+=2   # línea en blanco entre edificios

    # Total del proyecto
    tot(ws_e,r,1,'TOTAL PROYECTO',ha='left')
    tot(ws_e,r,2,int(len(df_e)),'0','center')
    tot(ws_e,r,3,float(df_e['GFA'].sum()),'#,##0.00')
    tot(ws_e,r,4,float(df_e['SV'].sum()),'#,##0.00')
    tot(ws_e,r,5,None)
    ws_e.row_dimensions[r].height=20
    ws_e.freeze_panes='A2'

    # ═══════════════════════════════════════════════════════
    # HOJA 6 — SUPERFICIES x PISO (por edificio: construido + vendible)
    # ═══════════════════════════════════════════════════════
    _hoja_superficies(wb, df)

    # ── Serializar a base64 ───────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode('ascii')


def superficies_xlsx(csv_text: str, n_sub: int, ediciones=None, extra=None) -> str:
    """Excel «tipo Mobil»: Superficies x Piso + (opcional) Normativa y Estacionamientos.

    `extra` (dict, opcional, lo arma el front con lo que el usuario ve en pantalla):
      { "normas": {...}, "estac": {...} }
    Devuelve base64 del .xlsx.
    """
    df, n_sub = _construir_df(csv_text, n_sub, ediciones)
    extra = extra or {}
    wb = Workbook()
    wb.remove(wb.active)  # quitar la hoja vacía por defecto
    _hoja_superficies(wb, df)
    if extra.get('normas'):
        _hoja_normativa(wb, extra['normas'])
    if extra.get('estac'):
        _hoja_estacionamientos(wb, extra['estac'])
    if not wb.sheetnames:  # sin nada → hoja mínima para no fallar
        ws = wb.create_sheet('Superficies x Piso')
        ws.cell(1, 1, 'Sin datos de superficies.')
    buf = io.BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode('ascii')


def normativa_xlsx(normas_data) -> str:
    """Excel «tipo Mobil» SÓLO con la hoja Normativa.

    `normas_data` = payload que arma el front (normasParaExcel):
      { "columnas": [{id,nombre}], "filas": [{label, grupo?, valores}] }
    No requiere CSV: funciona con los inputs/teaser que el usuario ve en pantalla.
    Devuelve base64 del .xlsx.
    """
    wb = Workbook()
    wb.remove(wb.active)
    if normas_data and (normas_data.get('filas')):
        _hoja_normativa(wb, normas_data)
    if not wb.sheetnames:
        ws = wb.create_sheet('Normativa')
        ws.cell(1, 1, 'Sin datos de normativa.')
    buf = io.BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode('ascii')


def _hoja_normativa(wb, data):
    """Hoja 'Normativa': parámetros × Fórmula × columnas (zonas/fusión/propuesto).
    data = { 'columnas': [{id,nombre}], 'filas': [{label, formula?, grupo?, valores:{colId: str}}] }
    Layout: A=Parámetro, B=Fórmula/factor, C…=columnas de datos."""
    cols = data.get('columnas') or []
    filas = data.get('filas') or []
    if not cols or not filas:
        return
    ws = wb.create_sheet('Normativa')
    ws.sheet_view.showGridLines = False
    DATA0 = 3                       # primera columna de datos (tras Parámetro y Fórmula)
    LAST = 2 + len(cols)
    ws.merge_cells(f'A1:{get_column_letter(LAST)}1')
    c = ws.cell(1, 1, 'NORMAS URBANÍSTICAS')
    c.fill = fill(C_HDR_BG); c.font = fnt(True, 13, C_HDR_FT); c.alignment = aln('center')
    ws.row_dimensions[1].height = 28
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    hdr(ws, 2, 1, 'Parámetro', ha='left')
    hdr(ws, 2, 2, 'Fórmula / factor', ha='left')
    for i, col in enumerate(cols, DATA0):
        hdr(ws, 2, i, str(col.get('nombre', '')))
        ws.column_dimensions[get_column_letter(i)].width = 18
    r = 3
    for f in filas:
        if f.get('grupo'):
            ws.merge_cells(f'A{r}:{get_column_letter(LAST)}{r}')
            cc = ws.cell(r, 1, str(f.get('label', '')))
            cc.fill = fill(C_TOTAL_BG); cc.font = fnt(True, 10, C_TOTAL_FT); cc.alignment = aln('left')
            r += 1
            continue
        ca = ws.cell(r, 1, str(f.get('label', ''))); ca.font = fnt(False, 10); ca.alignment = aln('left'); ca.border = brd()
        cf = ws.cell(r, 2, str(f.get('formula', '') or '') or None)
        cf.font = fnt(False, 9, '7A7A7A'); cf.alignment = aln('left'); cf.border = brd()
        vals = f.get('valores') or {}
        for i, col in enumerate(cols, DATA0):
            v = vals.get(col.get('id'), '')
            cc = ws.cell(r, i, v if v != '' else None)
            cc.font = fnt(False, 10); cc.alignment = aln('right'); cc.border = brd()
        r += 1


def _hoja_estacionamientos(wb, data):
    """Hoja 'Estacionamientos': Superficie y Subterráneo con m²/cajones/ratio/holgura.
    data = { 'bloques': [{titulo, gfa, cajones, ratio}] }"""
    bloques = data.get('bloques') or []
    if not bloques:
        return
    ws = wb.create_sheet('Estacionamientos')
    ws.sheet_view.showGridLines = False
    ws.merge_cells('A1:E1')
    c = ws.cell(1, 1, 'ESTACIONAMIENTOS — RATIO POR CAJÓN')
    c.fill = fill(C_HDR_BG); c.font = fnt(True, 13, C_HDR_FT); c.alignment = aln('center')
    ws.row_dimensions[1].height = 28
    ws.column_dimensions['A'].width = 24
    for col in 'BCDE':
        ws.column_dimensions[col].width = 16
    for i, h in enumerate(['Nivel', 'Superficie m²', 'Cajones', 'm²/cajón', 'Holgura'], 1):
        hdr(ws, 2, i, h, ha='left' if i == 1 else 'center')
    r = 3
    for b in bloques:
        gfa = float(b.get('gfa') or 0)
        caj = float(b.get('cajones') or 0)
        ratio = (gfa / caj) if caj > 0 else float(b.get('ratio') or 0)
        holg = 'Amplia' if ratio >= 30 else ('Justa' if ratio >= 25 else ('Ajustada' if ratio > 0 else '—'))
        ws.cell(r, 1, str(b.get('titulo', ''))).font = fnt(True, 10)
        cc = ws.cell(r, 2, round(gfa, 2) if gfa else None); cc.number_format = '#,##0.00'; cc.alignment = aln('right')
        cc = ws.cell(r, 3, int(caj) if caj else None); cc.number_format = '0'; cc.alignment = aln('center')
        cc = ws.cell(r, 4, round(ratio, 1) if ratio else None); cc.number_format = '0.0'; cc.alignment = aln('center')
        ws.cell(r, 5, holg).alignment = aln('center')
        r += 1


def _hoja_superficies(wb, df):
    """Hoja 'Superficies x Piso': una sección por edificio (+ Total proyecto),
    piso × función con totales bajo/sobre NTN, Construido y Vendible + ratio."""
    df = df.copy()
    df['Edificio'] = df['Edificio'].fillna('Sin edificio')
    funciones = [f for f in ORDEN_FUNC if f in df['Canonico'].unique()]
    if not funciones:
        return
    nombre_de = {}
    for h, nm in zip(df['Edificio'], df.get('EdificioNombre', df['Edificio'])):
        nombre_de.setdefault(str(h), str(nm))

    gfa_por_ed = df.groupby('Edificio')['GFA'].sum().sort_values(ascending=False)
    orden_ed = [e for e in gfa_por_ed.index if e != 'Sin edificio']
    if 'Sin edificio' in gfa_por_ed.index:
        orden_ed.append('Sin edificio')

    ws = wb.create_sheet('Superficies x Piso')
    LAST = 1 + len(funciones) + 2   # Piso + funciones + Construido + Vendible
    ws.merge_cells(f'A1:{get_column_letter(LAST)}1')
    c = ws.cell(1, 1, 'SUPERFICIES POR PISO — CONSTRUIDO Y VENDIBLE x EDIFICIO')
    c.fill = fill(C_HDR_BG); c.font = fnt(True, 13, C_HDR_FT); c.alignment = aln('center')
    ws.row_dimensions[1].height = 28
    ws.column_dimensions['A'].width = 16
    for i in range(2, LAST + 1):
        ws.column_dimensions[get_column_letter(i)].width = 13

    def bloque(sub, r, titulo):
        ws.merge_cells(f'A{r}:{get_column_letter(LAST)}{r}')
        cab = ws.cell(r, 1, titulo)
        cab.fill = fill(C_TOTAL_BG); cab.font = fnt(True, 11, C_TOTAL_FT); cab.alignment = aln('left')
        r += 1
        hdr(ws, r, 1, 'Piso', ha='left')
        for i, fn in enumerate(funciones, 2):
            hdr(ws, r, i, fn, wrap=True)
        hdr(ws, r, len(funciones) + 2, 'Construido')
        hdr(ws, r, len(funciones) + 3, 'Vendible')
        r += 1

        etqs = sorted(sub['Etiqueta'].unique().tolist(), key=sort_etq)
        piv = sub.pivot_table(index='Etiqueta', columns='Canonico', values='GFA',
                              aggfunc='sum', fill_value=0)
        for etq in etqs:
            f_sub = sub[sub['Etiqueta'] == etq]
            ws.cell(r, 1, etq).font = fnt(True, 10)
            for i, fn in enumerate(funciones, 2):
                v = float(piv.loc[etq, fn]) if (etq in piv.index and fn in piv.columns) else 0.0
                cc = ws.cell(r, i, round(v, 2) if v else None); cc.number_format = '#,##0.00'; cc.alignment = aln('right')
            cc = ws.cell(r, len(funciones) + 2, round(float(f_sub[f_sub['Integra']]['GFA'].sum()), 2)); cc.number_format = '#,##0.00'
            cc = ws.cell(r, len(funciones) + 3, round(float(f_sub['SV'].sum()), 2)); cc.number_format = '#,##0.00'
            r += 1

        tg = float(sub[sub['Integra']]['GFA'].sum())
        tv = float(sub['SV'].sum())
        tot(ws, r, 1, 'TOTAL', ha='left')
        for i, fn in enumerate(funciones, 2):
            v = float(sub[sub['Canonico'] == fn]['GFA'].sum())
            tot(ws, r, i, round(v, 2) if v else None, '#,##0.00')
        tot(ws, r, len(funciones) + 2, round(tg, 2), '#,##0.00')
        tot(ws, r, len(funciones) + 3, round(tv, 2), '#,##0.00')
        r += 1
        ratio = (tv / tg) if tg > 0 else 0.0
        rc = ws.cell(r, len(funciones) + 2, 'Vendible/Construido'); rc.font = fnt(False, 9, '555555'); rc.alignment = aln('right')
        rc = ws.cell(r, len(funciones) + 3, round(ratio, 4)); rc.number_format = '0.0%'; rc.font = fnt(True, 10)
        return r + 2

    r = 2
    r = bloque(df, r, 'TOTAL PROYECTO')
    for ed in orden_ed:
        r = bloque(df[df['Edificio'] == ed], r, nombre_de.get(str(ed), str(ed)))


# Matriz piso × función para graficar en la UI (barras horizontales apiladas)
def matriz_cabida(csv_text: str, n_sub: int, ediciones=None) -> dict:
    df, n_sub = _construir_df(csv_text, n_sub, ediciones)
    funciones = [f for f in ORDEN_FUNC if f in df['Canonico'].unique()]

    def sort_etq(e):
        if e.startswith('ST-'):
            try: return -1000 + int(e[3:])
            except: return -999
        elif e.startswith('N'):
            try: return int(e[1:])
            except: return 9999
        return 9998

    etiquetas = sorted(df['Etiqueta'].unique().tolist(), key=sort_etq)

    piv = df.pivot_table(index='Etiqueta', columns='Canonico', values='GFA',
                         aggfunc='sum', fill_value=0)
    datos = []
    for etq in etiquetas:
        fila = {'etiqueta': etq, 'es_sub': etq.startswith('ST-')}
        for fn in funciones:
            v = float(piv.loc[etq, fn]) if (etq in piv.index and fn in piv.columns) else 0.0
            fila[fn] = round(v, 2)
        datos.append(fila)

    return {
        'etiquetas': etiquetas,
        'funciones': funciones,
        'colores': {fn: '#' + COLOR_CANONICO.get(fn, 'BFBFBF') for fn in funciones},
        'datos': datos,
    }


# Cuadro de superficies: edificio × piso × función, con Construido y Vendible.
# Reemplaza/complementa la matriz: entrega la tabla "tipo Mobil" + datos del gráfico.
def cabida_superficies(csv_text: str, n_sub: int, ediciones=None) -> dict:
    """Superficies por edificio × piso × función (GFA construido + SV vendible).

    Devuelve, por cada edificio y por el TOTAL del proyecto (agregando por etiqueta):
      - pisos: filas con celdas por función, construido (GFA que integra) y vendible (SV).
      - subtotales bajo/sobre NTN, total y ratio vendible/construido.
    Reutiliza FACTOR_VENTA/CONSTRUIDO/GRUPOS/ORDEN_FUNC; no inventa superficies.
    """
    df, n_sub = _construir_df(csv_text, n_sub, ediciones)
    df = df.copy()
    df['Edificio'] = df['Edificio'].fillna('Sin edificio')

    # Funciones presentes, en orden canónico.
    funciones = [f for f in ORDEN_FUNC if f in df['Canonico'].unique()]

    # Estructura de columnas agrupadas (para la tabla). Orden de grupos del cuadro Mobil.
    ORDEN_GRUPOS = ['Otros usos', 'Residencial', 'Comercial', 'Oficinas']
    ETIQUETA_GRUPO = {
        'Otros usos': 'Estacionamientos', 'Residencial': 'Residencial',
        'Comercial': 'Eq. Comercio', 'Oficinas': 'Eq. Oficinas',
    }
    grupos_cols = []
    for g in ORDEN_GRUPOS:
        fns = [f for f in GRUPOS.get(g, []) if f in funciones]
        if not fns:
            continue
        grupos_cols.append({
            'grupo': ETIQUETA_GRUPO.get(g, g),
            'funciones': fns,
            'colores': {f: '#' + COLOR_CANONICO.get(f, 'BFBFBF') for f in fns},
        })

    # Mapa hash → nombre legible (lo trae _construir_df).
    nombre_de = {}
    for h, nm in zip(df['Edificio'], df['EdificioNombre']):
        nombre_de.setdefault(str(h), str(nm))

    # Orden de edificios por GFA (mayor a menor), 'Sin edificio' al final.
    gfa_por_ed = df.groupby('Edificio')['GFA'].sum().sort_values(ascending=False)
    orden_ed = [e for e in gfa_por_ed.index if e != 'Sin edificio']
    if 'Sin edificio' in gfa_por_ed.index:
        orden_ed.append('Sin edificio')

    def construir_bloque(sub):
        """Bloque {pisos, subtotal_bajo, subtotal_sobre, total, ratio} para un df.
        Orden de la tabla: pisos altos arriba (N alto → N1) y subterráneos abajo
        empezando por el menos profundo (ST-1, ST-2, …)."""
        etiquetas = sorted(sub['Etiqueta'].unique().tolist(), key=sort_etq_tabla)
        piv_gfa = sub.pivot_table(index='Etiqueta', columns='Canonico',
                                  values='GFA', aggfunc='sum', fill_value=0)
        # Aporte manual (áreas agregadas) por celda, para distinguirlo del CSV.
        es_manual = sub['Id'].astype(str).str.startswith('manual')
        sub_man = sub[es_manual]
        piv_man = (sub_man.pivot_table(index='Etiqueta', columns='Canonico',
                                       values='GFA', aggfunc='sum', fill_value=0)
                   if len(sub_man) else None)
        pisos = []
        for etq in etiquetas:
            celdas = {}
            celdas_manual = {}
            for fn in funciones:
                v = float(piv_gfa.loc[etq, fn]) if (etq in piv_gfa.index and fn in piv_gfa.columns) else 0.0
                celdas[fn] = round(v, 2)
                m = (float(piv_man.loc[etq, fn]) if (piv_man is not None and etq in piv_man.index
                                                     and fn in piv_man.columns) else 0.0)
                if m:
                    celdas_manual[fn] = round(m, 2)
            f_sub = sub[sub['Etiqueta'] == etq]
            construido = float(f_sub[f_sub['Integra']]['GFA'].sum())
            vendible = float(f_sub['SV'].sum())
            pisos.append({
                'etiqueta': etq,
                'es_sub': etq.startswith('ST-'),
                'celdas': celdas,
                'celdas_manual': celdas_manual,
                'construido': round(construido, 2),
                'vendible': round(vendible, 2),
                'total': round(float(f_sub['GFA'].sum()), 2),
            })

        def agregar(filas):
            celdas = {fn: round(sum(p['celdas'][fn] for p in filas), 2) for fn in funciones}
            return {
                'celdas': celdas,
                'construido': round(sum(p['construido'] for p in filas), 2),
                'vendible': round(sum(p['vendible'] for p in filas), 2),
                'total': round(sum(p['total'] for p in filas), 2),
            }

        bajo = [p for p in pisos if p['es_sub']]
        sobre = [p for p in pisos if not p['es_sub']]
        total = agregar(pisos)
        ratio = (total['vendible'] / total['construido']) if total['construido'] > 0 else 0.0
        return {
            'pisos': pisos,
            'subtotal_bajo': agregar(bajo) if bajo else None,
            'subtotal_sobre': agregar(sobre) if sobre else None,
            'total': total,
            'ratio': round(ratio, 4),
        }

    por_edificio = {}
    lista_eds = []
    for ed in orden_ed:
        por_edificio[str(ed)] = construir_bloque(df[df['Edificio'] == ed])
        lista_eds.append({'id': str(ed), 'nombre': nombre_de.get(str(ed), str(ed))})
    # Total proyecto (agrega por etiqueta, no por edificio).
    por_edificio['__total__'] = construir_bloque(df)

    return {
        'edificios': lista_eds,           # para el selector (+ "Total" en el front)
        'funciones': funciones,
        'colores': {fn: '#' + COLOR_CANONICO.get(fn, 'BFBFBF') for fn in funciones},
        'grupos': grupos_cols,            # estructura de columnas Útil/Común por grupo
        'por_edificio': por_edificio,
    }


def metricas_normativas(csv_text: str, n_sub: int, ediciones=None) -> dict:
    """Métricas de la cabida que alimentan la tabla de Normas Urbanísticas.

    - constructibilidad: Σ SOBRE NTN de Residencial Útil + Comercio (GFA) + Oficinas (GFA).
    - ocupacion_p1: GFA total del piso N1.
    - viviendas: nº de departamentos (unidades habitacionales).
    """
    df, n_sub = _construir_df(csv_text, n_sub, ediciones)
    sobre = df[~df['Etiqueta'].astype(str).str.startswith('ST-')]

    res_util = float(sobre[sobre['Canonico'] == 'Residencial Util']['GFA'].sum())
    comercio = float(sobre[sobre['Canonico'].isin(['Comercial Util', 'Comercial Comun'])]['GFA'].sum())
    oficinas = float(sobre[sobre['Canonico'].isin(['Oficinas Util', 'Oficinas Comun'])]['GFA'].sum())
    constructibilidad = res_util + comercio + oficinas

    ocupacion_p1 = float(df[df['Etiqueta'] == 'N1']['GFA'].sum())
    viviendas = int(df['Type'].apply(es_departamento).sum())

    return {
        'constructibilidad': round(constructibilidad, 2),
        'ocupacion_p1': round(ocupacion_p1, 2),
        'viviendas': viviendas,
        'desglose': {
            'residencial_util': round(res_util, 2),
            'comercio_gfa': round(comercio, 2),
            'oficinas_gfa': round(oficinas, 2),
        },
    }


# Resumen rápido para mostrar en la UI (sin generar el Excel completo)
def resumen_cabida(csv_text: str, n_sub: int, ediciones=None) -> dict:
    df, n_sub = _construir_df(csv_text, n_sub, ediciones)
    venta = float(df['SV'].sum())
    const = float(df[df['Integra']]['GFA'].sum())
    efic = venta / const if const > 0 else 0
    return {
        'elementos': int(len(df)),
        'venta': round(venta),
        'construido': round(const),
        'eficiencia': round(efic, 4),
    }
